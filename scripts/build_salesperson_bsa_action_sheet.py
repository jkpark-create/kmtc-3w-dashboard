from __future__ import annotations

import json
import math
import numbers
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "output"
TEAM_FILTER = "OBT"
MONTHS = [f"2025{m:02d}" for m in range(1, 13)] + [f"2026{m:02d}" for m in range(1, 7)]
MONTH_LABELS = {m: f"{m[:4]}-{m[4:]}" for m in MONTHS}
Q2_PROGRESS_KEY = "2026_Q2_W14_19"
Q2_PROGRESS_LABEL = "2026 Q2 W14-19"
Q2_PROGRESS_START = datetime(2026, 4, 5)
Q2_PROGRESS_END = datetime(2026, 5, 10)
Q2_PROGRESS_WEEKS = set(range(14, 20))
MISSING = "(미지정)"
NO_BASIS_LABEL = "(no 2025 basis)"
NO_BASIS_LEVEL = "no 2025 basis"
NO_BASIS_NOTE = (
    "2025년 같은 POR_port + DLY_port route에 Normal LST_TEU 배분 기준이 없어 "
    "해당 route BSA를 담당자별로 나누지 못하고 예외로 남긴 값입니다."
)
ALWAYS_KEEP_SALES = {MISSING, NO_BASIS_LABEL}
SALES_SCOPED_TAB_PREFIXES = ("CN_",)
CN_NKG_PORTS = frozenset({
    "AIA", "AQG", "CGD", "CGS", "CKG", "CKQ", "CSX", "CZH",
    "CZX", "FLG", "HFE", "HSI", "JIA", "JIN", "JJG", "LUZ",
    "MSN", "NCH", "NKG", "NTG", "TAZ", "TCG", "TOL", "WHI",
    "WUH", "WUW", "WZH", "YCH", "YYA", "YZH", "YZR", "ZHE", "ZJG",
})


def latest_dataset_id() -> str:
    candidates: list[tuple[float, str]] = []
    for path in OUT_DIR.glob("_cache_*.parquet"):
        match = re.fullmatch(r"_cache_(\d{8})", path.stem)
        if match:
            candidates.append((path.stat().st_mtime, match.group(1)))
    if candidates:
        return max(candidates)[1]
    return "20260508"


CURRENT_DATASET_ID = os.environ.get("SALES_BSA_DATASET_ID", latest_dataset_id())

PERIODS: list[tuple[str, str, list[str]]] = [
    ("202501", "2025-01", ["202501"]),
    ("202502", "2025-02", ["202502"]),
    ("202503", "2025-03", ["202503"]),
    ("2025_Q1", "2025 Q1", ["202501", "202502", "202503"]),
    ("202504", "2025-04", ["202504"]),
    ("202505", "2025-05", ["202505"]),
    ("202506", "2025-06", ["202506"]),
    ("2025_Q2", "2025 Q2", ["202504", "202505", "202506"]),
    ("202507", "2025-07", ["202507"]),
    ("202508", "2025-08", ["202508"]),
    ("202509", "2025-09", ["202509"]),
    ("2025_Q3", "2025 Q3", ["202507", "202508", "202509"]),
    ("202510", "2025-10", ["202510"]),
    ("202511", "2025-11", ["202511"]),
    ("202512", "2025-12", ["202512"]),
    ("2025_Q4", "2025 Q4", ["202510", "202511", "202512"]),
    ("2025_Total", "2025 Total", [f"2025{m:02d}" for m in range(1, 13)]),
    ("202601", "2026-01", ["202601"]),
    ("202602", "2026-02", ["202602"]),
    ("202603", "2026-03", ["202603"]),
    ("2026_Q1", "2026 Q1", ["202601", "202602", "202603"]),
    (Q2_PROGRESS_KEY, Q2_PROGRESS_LABEL, [Q2_PROGRESS_KEY]),
]


def norm_text(value: object, fallback: str = "") -> str:
    if pd.isna(value):
        return fallback
    text = str(value).strip()
    return fallback if text.lower() in {"", "nan", "none", "nat"} else text


def period_kind(period_key: str) -> str:
    if period_key.endswith("_Total"):
        return "year"
    if "_Q" in period_key:
        return "quarter"
    return "month"


def is_no_basis_value(value: object) -> bool:
    return norm_text(value) in {NO_BASIS_LABEL, NO_BASIS_LEVEL}


def to_number(series: pd.Series) -> pd.Series:
    return pd.to_numeric(
        series.fillna("").astype(str).str.replace(",", "", regex=False).str.strip(),
        errors="coerce",
    ).fillna(0.0)


def excel_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, numbers.Integral):
        return int(value)
    if isinstance(value, numbers.Real):
        number = float(value)
        if math.isnan(number) or math.isinf(number):
            return None
        return int(number) if number.is_integer() else number
    return value


def classify_team(origin: object, dest: object) -> str:
    o = norm_text(origin)
    d = norm_text(dest)
    if o not in ("KR", "JP") and d != "KR":
        return "OBT"
    if o == "KR" and d != "JP":
        return "EST"
    if o != "JP" and d == "KR":
        return "IST"
    return "JBT"


def tab_key(origin: object, ori_port: object) -> str:
    origin = norm_text(origin, "UNKNOWN")
    port = norm_text(ori_port, "UNKNOWN")
    if origin == "CN":
        if port in CN_NKG_PORTS:
            return "CN_NKG"
        return "CN_SHK_DCB" if port in {"SHK", "DCB"} else f"CN_{port}"
    if origin == "VN":
        if port in {"SGN", "CMP"}:
            return "VN_SGN_CMP"
        if port == "HPH":
            return "VN_HPH"
        return f"VN_{port}"
    if origin == "ID":
        if port == "JKT":
            return "JKT"
        if port == "SUB":
            return "SUB"
        return "ID_out"
    if origin == "MY":
        if port in {"PKG", "PKW"}:
            return "PKG+PKW"
        if port == "PEN":
            return "PEN"
        if port == "PGU":
            return "PGU"
        return "MY_out"
    return origin


def parse_week_start(value: object) -> datetime | None:
    text = norm_text(value)
    if not text:
        return None
    match = re.match(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", text)
    if not match:
        return None
    year, month, day = (int(part) for part in match.groups())
    try:
        return datetime(year, month, day)
    except ValueError:
        return None


def is_q2_progress_week(value: object) -> bool:
    week_start = parse_week_start(value)
    return bool(week_start and Q2_PROGRESS_START <= week_start <= Q2_PROGRESS_END)


def compute_route_profit_type(df: pd.DataFrame) -> pd.Series:
    """Classify each shipper-route as high/low profit by POR_port + DLY_port CM1/TEU."""
    out = pd.Series([""] * len(df), index=df.index, dtype="object")
    if df.empty:
        return out

    valid = df.loc[
        df["status"].eq("Normal")
        & df["cm1"].ne(0)
        & df["lst"].gt(0)
        & df["shipper_code"].ne("")
        & df["ori_port"].ne("")
        & df["dst_port"].ne("")
    ].copy()
    if valid.empty:
        return out

    route = (
        valid.groupby(["ori_port", "dst_port"], dropna=False)
        .agg(route_cm1=("cm1", "sum"), route_teu=("lst", "sum"))
        .reset_index()
    )
    route["route_cm1_teu"] = route["route_cm1"] / route["route_teu"]

    shipper_route = (
        valid.groupby(["shipper_code", "ori_port", "dst_port"], dropna=False)
        .agg(shipper_cm1=("cm1", "sum"), shipper_teu=("lst", "sum"))
        .reset_index()
    )
    shipper_route["shipper_cm1_teu"] = shipper_route["shipper_cm1"] / shipper_route["shipper_teu"]
    shipper_route = shipper_route.merge(
        route[["ori_port", "dst_port", "route_cm1_teu"]],
        on=["ori_port", "dst_port"],
        how="left",
    )
    shipper_route["profit_type"] = shipper_route.apply(
        lambda row: "고수익" if row["shipper_cm1_teu"] >= row["route_cm1_teu"] else "저수익",
        axis=1,
    )
    lookup = {
        (row.shipper_code, row.ori_port, row.dst_port): row.profit_type
        for row in shipper_route.itertuples(index=False)
    }
    out.loc[:] = [
        lookup.get((shipper, ori_port, dst_port), "")
        for shipper, ori_port, dst_port in zip(df["shipper_code"], df["ori_port"], df["dst_port"])
    ]
    return out


def load_booking() -> pd.DataFrame:
    cache_2025 = OUT_DIR / "_cache_2025.parquet"
    cache_current = OUT_DIR / f"_cache_{CURRENT_DATASET_ID}.parquet"
    if not cache_2025.exists():
        raise FileNotFoundError(f"Missing 2025 cache: {cache_2025}")
    if not cache_current.exists():
        raise FileNotFoundError(f"Missing current cache: {cache_current}")

    columns = [
        "BKG_NO",
        "BKG_SHPR_CST_NO",
        "BKG_SHPR_CST_ENM",
        "POR_CTR_CD",
        "POR_PLC_CD",
        "DLY_CTR_CD",
        "DLY_PLC_CD",
        "FST_TEU",
        "LST_Status",
        "LST_TEU",
        "CM1",
        "YYYYMM",
        "week_start_date",
        "Booking_date",
        "Lead_time (BKG_Sche)",
        "Salesman_POR",
        "고/저",
    ]
    frames = []
    for path in (cache_2025, cache_current):
        frame = pd.read_parquet(path)
        # Booking_date may be absent in older caches; provide an empty fallback.
        missing = [c for c in columns if c not in frame.columns]
        for c in missing:
            frame[c] = ""
        frame = frame[columns].rename(
            columns={
            "BKG_SHPR_CST_NO": "shipper_code",
            "BKG_SHPR_CST_ENM": "shipper_name",
            "POR_CTR_CD": "origin",
            "POR_PLC_CD": "ori_port",
            "DLY_CTR_CD": "dest",
            "DLY_PLC_CD": "dst_port",
            "FST_TEU": "fst",
            "LST_Status": "status",
            "LST_TEU": "lst",
            "CM1": "cm1",
            "YYYYMM": "yyyymm",
            "week_start_date": "week_start",
            "Booking_date": "booking_date",
            "Lead_time (BKG_Sche)": "lead_time",
            "Salesman_POR": "sales",
            "고/저": "profit_type",
        }
        )
        for col in [
            "shipper_code",
            "shipper_name",
            "origin",
            "ori_port",
            "dest",
            "dst_port",
            "status",
            "yyyymm",
            "week_start",
            "lead_time",
            "sales",
            "profit_type",
        ]:
            frame[col] = frame[col].map(norm_text)
        frame["fst"] = to_number(frame["fst"])
        frame["lst"] = to_number(frame["lst"])
        frame["cm1"] = to_number(frame["cm1"])
        frame["profit_type"] = compute_route_profit_type(frame)
        frame["_q2_progress"] = frame["week_start"].map(is_q2_progress_week)
        frames.append(frame)
    df = pd.concat(frames, ignore_index=True)
    df["sales"] = df["sales"].replace("", MISSING)
    df["team"] = [classify_team(o, d) for o, d in zip(df["origin"], df["dest"])]
    df["tab"] = [tab_key(o, p) for o, p in zip(df["origin"], df["ori_port"])]
    df = df.loc[df["yyyymm"].isin(MONTHS) & df["team"].eq(TEAM_FILTER)].copy()
    df = _remap_2025_sales_to_current_owner(df)
    return df


# 1Q 2026 booking 중 가장 최근 booking_date의 Salesman_POR을 화주별 "현재 담당자"로 채택하고,
# 2025 booking 레코드의 sales를 그 매핑 값으로 덮어씁니다. 1Q 2026에 booking이 없는 화주는
# 2025 담당자를 그대로 유지합니다 (사용자 정책).
Q1_2026_MONTHS = frozenset({"202601", "202602", "202603"})


def _remap_2025_sales_to_current_owner(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    q1 = df.loc[df["yyyymm"].isin(Q1_2026_MONTHS) & df["shipper_code"].ne("")].copy()
    if q1.empty:
        return df
    q1["_bdate"] = pd.to_datetime(q1["booking_date"], errors="coerce")
    # If booking_date is missing for all Q1 rows, fall back to yyyymm-only ordering so the
    # mapping still resolves to *some* deterministic Salesman_POR per shipper.
    if q1["_bdate"].isna().all():
        q1 = q1.sort_values(["yyyymm"], ascending=False)
    else:
        # Latest booking_date first; rows with no parseable date sink to the bottom.
        q1["_bdate_filled"] = q1["_bdate"].fillna(pd.Timestamp.min)
        q1 = q1.sort_values(["_bdate_filled"], ascending=False)
    # The MISSING placeholder shouldn't claim ownership unless it's all we have for the shipper.
    q1 = q1.assign(_priority=q1["sales"].eq(MISSING).astype(int))
    q1 = q1.sort_values(["_priority"], ascending=True, kind="stable")
    current_owner = q1.drop_duplicates("shipper_code", keep="first").set_index("shipper_code")["sales"]
    is_2025 = df["yyyymm"].str.startswith("2025") & df["shipper_code"].ne("")
    mapped = df.loc[is_2025, "shipper_code"].map(current_owner)
    df.loc[is_2025, "sales"] = mapped.fillna(df.loc[is_2025, "sales"])
    return df


def load_bsa() -> pd.DataFrame:
    path = OUT_DIR / f"BSA_raw_monthly3W_{CURRENT_DATASET_ID}.csv"
    if not path.exists():
        raise FileNotFoundError(f"Missing BSA file: {path}")
    bsa = pd.read_csv(path, dtype=str, encoding="utf-8-sig")
    bsa = bsa.rename(
        columns={
            "POR_Country": "origin",
            "POR_PORT": "ori_port",
            "DLY_Country": "dest",
            "DLY_PORT": "dst_port",
            "YYYYMM": "yyyymm",
            "TEU_BSA (Actual)": "route_bsa",
            "WW": "ww",
        }
    )
    for col in ["origin", "ori_port", "dest", "dst_port", "yyyymm", "ww"]:
        bsa[col] = bsa[col].map(norm_text)
    bsa["team"] = bsa.get("team", bsa.get("Sales Team", "")).map(norm_text).str.upper()
    bsa["route_bsa"] = to_number(bsa["route_bsa"])
    bsa["ww_num"] = pd.to_numeric(bsa["ww"], errors="coerce").fillna(0).astype(int)
    bsa["tab"] = [tab_key(o, p) for o, p in zip(bsa["origin"], bsa["ori_port"])]
    bsa = bsa.loc[
        bsa["yyyymm"].isin(MONTHS) & bsa["team"].eq(TEAM_FILTER) & bsa["route_bsa"].gt(0)
    ].copy()
    keys = ["yyyymm", "tab", "team", "origin", "ori_port", "dest", "dst_port"]
    monthly = bsa.groupby(keys, dropna=False)["route_bsa"].sum().reset_index()
    q2 = bsa.loc[bsa["yyyymm"].str.startswith("2026") & bsa["ww_num"].isin(Q2_PROGRESS_WEEKS)].copy()
    if q2.empty:
        return monthly
    q2["yyyymm"] = Q2_PROGRESS_KEY
    q2 = q2.groupby(keys, dropna=False)["route_bsa"].sum().reset_index()
    return pd.concat([monthly, q2], ignore_index=True)


def metric_sum(df: pd.DataFrame, mask: pd.Series, value_col: str) -> pd.DataFrame:
    keys = ["tab", "team", "sales", "yyyymm"]
    out = (
        df.loc[mask, keys + [value_col]]
        .groupby(keys, dropna=False)[value_col]
        .sum()
        .reset_index()
        .rename(columns={value_col: "value"})
    )
    if "_q2_progress" in df.columns:
        q2_mask = mask & df["_q2_progress"].fillna(False)
        if q2_mask.any():
            q2 = df.loc[q2_mask, keys + [value_col]].copy()
            q2["yyyymm"] = Q2_PROGRESS_KEY
            q2 = (
                q2.groupby(keys, dropna=False)[value_col]
                .sum()
                .reset_index()
                .rename(columns={value_col: "value"})
            )
            out = pd.concat([out, q2], ignore_index=True)
    return out


def build_basis(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, tuple[list[str], dict[tuple, pd.DataFrame]]]]:
    basis_rows = df.loc[
        df["yyyymm"].str.startswith("2025") & df["status"].eq("Normal") & df["lst"].gt(0)
    ].copy()

    basis_detail = (
        basis_rows.groupby(
            [
                "tab",
                "team",
                "origin",
                "ori_port",
                "dest",
                "dst_port",
                "shipper_code",
                "shipper_name",
                "sales",
            ],
            dropna=False,
        )["lst"]
        .sum()
        .reset_index()
        .rename(columns={"lst": "2025_LST_TEU"})
    )

    levels = {
        "same target tab + DLY_port": ["team", "tab", "dest", "dst_port"],
    }
    lookups: dict[str, tuple[list[str], dict[tuple, pd.DataFrame]]] = {}
    for level, keys in levels.items():
        grouped = (
            basis_rows.groupby(keys + ["sales"], dropna=False)["lst"]
            .sum()
            .reset_index()
            .rename(columns={"lst": "basis_lST_TEU"})
        )
        lookup: dict[tuple, pd.DataFrame] = {}
        for key, group in grouped.groupby(keys, dropna=False):
            key_tuple = key if isinstance(key, tuple) else (key,)
            lookup[key_tuple] = group[["sales", "basis_lST_TEU"]].reset_index(drop=True)
        lookups[level] = (keys, lookup)
    return basis_detail, lookups


def allocate_bsa(bsa: pd.DataFrame, lookups: dict[str, tuple[list[str], dict[tuple, pd.DataFrame]]]) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    route_cols = ["yyyymm", "tab", "team", "origin", "ori_port", "dest", "dst_port"]
    for route in bsa.itertuples(index=False):
        route_dict = route._asdict()
        selected_level = NO_BASIS_LEVEL
        selected_keys: list[str] = []
        selected = pd.DataFrame()
        selected_total = 0.0
        for level, (keys, lookup) in lookups.items():
            key = tuple(route_dict[k] for k in keys)
            group = lookup.get(key)
            total = float(group["basis_lST_TEU"].sum()) if group is not None else 0.0
            if total > 0:
                selected_level = level
                selected_keys = keys
                selected = group
                selected_total = total
                break

        route_bsa = float(route_dict["route_bsa"])
        if selected_total <= 0:
            base = {col: route_dict[col] for col in route_cols}
            rows.append(
                {
                    **base,
                    "sales": NO_BASIS_LABEL,
                    "basis_lST_TEU": 0.0,
                    "basis_total_lST_TEU": 0.0,
                    "allocation_share": 0.0,
                    "route_bsa": route_bsa,
                    "allocated_bsa": route_bsa,
                    "allocation_level": selected_level,
                    "allocation_key": "",
                }
            )
            continue

        allocation_key = ", ".join(f"{k}={route_dict[k]}" for k in selected_keys)
        for item in selected.itertuples(index=False):
            basis = float(item.basis_lST_TEU)
            share = basis / selected_total
            base = {col: route_dict[col] for col in route_cols}
            rows.append(
                {
                    **base,
                    "sales": item.sales,
                    "basis_lST_TEU": basis,
                    "basis_total_lST_TEU": selected_total,
                    "allocation_share": share,
                    "route_bsa": route_bsa,
                    "allocated_bsa": route_bsa * share,
                    "allocation_level": selected_level,
                    "allocation_key": allocation_key,
                }
            )

    return pd.DataFrame(rows)


def sales_metric_from_allocation(allocation: pd.DataFrame) -> pd.DataFrame:
    keys = ["tab", "team", "sales", "yyyymm"]
    return (
        allocation.groupby(keys, dropna=False)["allocated_bsa"]
        .sum()
        .reset_index()
        .rename(columns={"allocated_bsa": "value"})
    )


def pivot_sum(long_df: pd.DataFrame, tab: str, rows: pd.DataFrame) -> pd.DataFrame:
    frame = long_df.loc[long_df["tab"].eq(tab)].copy()
    if frame.empty:
        pivot = pd.DataFrame(index=pd.MultiIndex.from_frame(rows[["team", "sales"]]))
    else:
        pivot = frame.pivot_table(
            index=["team", "sales"],
            columns="yyyymm",
            values="value",
            aggfunc="sum",
            fill_value=0.0,
        )
    pivot = pivot.reindex(pd.MultiIndex.from_frame(rows[["team", "sales"]]), fill_value=0.0)
    required_members = set(MONTHS)
    for _, _, members in PERIODS:
        required_members.update(members)
    for member in required_members:
        if member not in pivot.columns:
            pivot[member] = 0.0
    out = rows.copy()
    for key, _, members in PERIODS:
        out[key] = pivot[members].sum(axis=1).to_numpy()
    return out


def pivot_ratio(numerator: pd.DataFrame, denominator: pd.DataFrame, tab: str, rows: pd.DataFrame) -> pd.DataFrame:
    n = pivot_sum(numerator, tab, rows)
    d = pivot_sum(denominator, tab, rows)
    out = rows.copy()
    for key, _, _ in PERIODS:
        den = d[key].replace(0, pd.NA)
        out[key] = n[key] / den
    return out


def sales_scope_enabled(tab: str) -> bool:
    return tab.startswith(SALES_SCOPED_TAB_PREFIXES)


def primary_sales_tab(metric_frames: Iterable[pd.DataFrame]) -> dict[str, str]:
    pieces = []
    for frame in metric_frames:
        if frame.empty or not {"tab", "sales", "value"}.issubset(frame.columns):
            continue
        item = frame.loc[:, ["tab", "sales", "value"]].copy()
        item["value"] = pd.to_numeric(item["value"], errors="coerce").fillna(0.0)
        item = item.loc[item["value"].gt(0) & item["sales"].notna()]
        if not item.empty:
            pieces.append(item)
    if not pieces:
        return {}
    totals = (
        pd.concat(pieces, ignore_index=True)
        .groupby(["sales", "tab"], dropna=False)["value"]
        .sum()
        .reset_index()
        .sort_values(["sales", "value", "tab"], ascending=[True, False, True])
    )
    return totals.drop_duplicates("sales").set_index("sales")["tab"].to_dict()


def apply_sales_scope(tab: str, rows: pd.DataFrame, metric_frames: Iterable[pd.DataFrame]) -> pd.DataFrame:
    if rows.empty or not sales_scope_enabled(tab):
        return rows
    primary = primary_sales_tab(metric_frames)
    keep = rows["sales"].isin(ALWAYS_KEEP_SALES) | rows["sales"].map(primary).eq(tab)
    return rows.loc[keep].reset_index(drop=True)


def restrict_metric_to_rows(frame: pd.DataFrame, rows: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame.copy()
    allowed = rows[["team", "sales"]].drop_duplicates()
    if allowed.empty:
        return frame.iloc[0:0].copy()
    return frame.merge(allowed, on=["team", "sales"], how="inner")


def restrict_source_to_rows(
    source: pd.DataFrame | tuple[pd.DataFrame, pd.DataFrame],
    rows: pd.DataFrame,
) -> pd.DataFrame | tuple[pd.DataFrame, pd.DataFrame]:
    if isinstance(source, tuple):
        return tuple(restrict_metric_to_rows(frame, rows) for frame in source)  # type: ignore[return-value]
    return restrict_metric_to_rows(source, rows)


def build_rows_for_tab(tab: str, metric_frames: Iterable[pd.DataFrame], sort_basis: pd.DataFrame) -> pd.DataFrame:
    metric_frames = list(metric_frames)
    pieces = []
    for frame in metric_frames:
        item = frame.loc[frame["tab"].eq(tab), ["team", "sales"]].drop_duplicates()
        if not item.empty:
            pieces.append(item)
    if not pieces:
        return pd.DataFrame(columns=["team", "sales"])
    rows = pd.concat(pieces, ignore_index=True).drop_duplicates()
    sort_values = pivot_sum(sort_basis, tab, rows)
    rows = rows.merge(
        sort_values[["team", "sales", "2025_Total", "2026_Q1"]],
        on=["team", "sales"],
        how="left",
    )
    rows = rows.sort_values(
        ["2025_Total", "2026_Q1", "team", "sales"],
        ascending=[False, False, True, True],
    )
    rows = rows[["team", "sales"]].reset_index(drop=True)
    return apply_sales_scope(tab, rows, metric_frames)


def add_total_row(table: pd.DataFrame, source: pd.DataFrame | tuple[pd.DataFrame, pd.DataFrame], tab: str, ratio: bool) -> pd.DataFrame:
    total = {"team": "TOTAL", "sales": "TOTAL"}

    def total_for(frame: pd.DataFrame, members: list[str]) -> float:
        matched = frame.loc[frame["tab"].eq(tab) & frame["yyyymm"].isin(members), "value"]
        return float(matched.sum()) if not matched.empty else 0.0

    if ratio:
        numerator, denominator = source
        for key, _, members in PERIODS:
            den = total_for(denominator, members)
            total[key] = None if den == 0 else total_for(numerator, members) / den
    else:
        for key, _, members in PERIODS:
            total[key] = total_for(source, members)  # type: ignore[arg-type]
    return pd.concat([pd.DataFrame([total]), table], ignore_index=True)


def safe_sheet_title(name: str, existing: set[str]) -> str:
    cleaned = re.sub(r"[:\\/?*\[\]]", "_", name)[:31] or "Sheet"
    candidate = cleaned
    i = 2
    while candidate in existing:
        suffix = f"_{i}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        i += 1
    existing.add(candidate)
    return candidate


def write_table(ws, start_row: int, title: str, table: pd.DataFrame, is_ratio: bool) -> int:
    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    quarter_header_fill = PatternFill("solid", fgColor="C6E0B4")
    quarter_value_fill = PatternFill("solid", fgColor="E2F0D9")
    year_header_fill = PatternFill("solid", fgColor="F4B183")
    year_value_fill = PatternFill("solid", fgColor="FCE4D6")
    no_basis_fill = PatternFill("solid", fgColor="F8CBAD")
    highlight_side = Side(style="thin", color="A6A6A6")
    highlight_border = Border(left=highlight_side, right=highlight_side)
    period_columns = {
        col: kind
        for col, (key, _, _) in enumerate(PERIODS, start=3)
        if (kind := period_kind(key)) != "month"
    }
    max_col = 2 + len(PERIODS)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_col)
    title_cell = ws.cell(start_row, 1, title)
    title_cell.font = Font(bold=True, color="FFFFFF")
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="left")

    headers = ["Team", "Salesman"] + [label for _, label, _ in PERIODS]
    header_row = start_row + 1
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        kind = period_columns.get(col)
        if kind == "quarter":
            cell.fill = quarter_header_fill
            cell.border = highlight_border
        elif kind == "year":
            cell.fill = year_header_fill
            cell.border = highlight_border
        elif header == "Salesman":
            cell.comment = Comment(NO_BASIS_NOTE, "Codex")

    data_row = header_row + 1
    for r_idx, row in enumerate(table.itertuples(index=False), start=data_row):
        row_values = list(row)
        for col, value in enumerate(row_values, start=1):
            value = excel_value(value)
            cell = ws.cell(r_idx, col, value)
            if is_no_basis_value(value):
                cell.fill = no_basis_fill
                cell.comment = Comment(NO_BASIS_NOTE, "Codex")
            if r_idx == data_row:
                cell.font = Font(bold=True)
                cell.fill = total_fill
            if col >= 3:
                if is_ratio:
                    cell.number_format = "0.0%"
                else:
                    cell.number_format = "#,##0.0"
                cell.alignment = Alignment(horizontal="right")
                kind = period_columns.get(col)
                if kind == "quarter":
                    cell.fill = quarter_value_fill
                    cell.border = highlight_border
                elif kind == "year":
                    cell.fill = year_value_fill
                    cell.border = highlight_border

    return data_row + len(table) + 1


def write_report_sheet(
    wb: Workbook,
    title: str,
    rows: pd.DataFrame,
    metrics: dict[str, pd.DataFrame | tuple[pd.DataFrame, pd.DataFrame]],
) -> None:
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    current_row = 1
    table_specs = [
        ("1. 전체 실적 LST_TEU (Normal 기준)", "lst", False),
        ("2. 2025 실적비중 기준 배분 BSA", "bsa", False),
        ("3. 3주전 부킹 FST_TEU (대시보드 WOS-3 기준)", "w3", False),
        ("4. 3주전 부킹 실선적 LST_TEU (Normal 기준)", "w3_norm_lst", False),
        ("5. 3주전 부킹 실선적률", "w3_ship_rate", True),
        ("6. 3주전 부킹 / BSA", "w3_bsa", True),
        ("7. 3주전 고수익 부킹 FST_TEU", "hi_w3", False),
        ("8. 3주전 고수익 부킹 비중", "hi_share", True),
    ]
    for table_title, key, is_ratio in table_specs:
        source = restrict_source_to_rows(metrics[key], rows)
        if is_ratio:
            numerator, denominator = source  # type: ignore[misc]
            table = pivot_ratio(numerator, denominator, title, rows)
            table = add_total_row(table, (numerator, denominator), title, ratio=True)
        else:
            table = pivot_sum(source, title, rows)  # type: ignore[arg-type]
            table = add_total_row(table, source, title, ratio=False)  # type: ignore[arg-type]
        current_row = write_table(ws, current_row, table_title, table, is_ratio)

    ws.freeze_panes = "A3"
    widths = [10, 18] + [11] * len(PERIODS)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_dataframe_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    no_basis_header_fill = PatternFill("solid", fgColor="F4B183")
    no_basis_fill = PatternFill("solid", fgColor="FCE4D6")
    for col, header in enumerate(df.columns, start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        if header in {"Salesman", "Allocation_Level"}:
            cell.comment = Comment(NO_BASIS_NOTE, "Codex")
        if header in {"Allocation_Level", "No_Basis_Rows", "No_Basis_BSA"}:
            cell.fill = no_basis_header_fill
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            value = excel_value(value)
            cell = ws.cell(r_idx, c_idx, value)
            if is_no_basis_value(value):
                cell.fill = no_basis_fill
                cell.comment = Comment(NO_BASIS_NOTE, "Codex")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(max(10, len(str(header)) + 2), 34)


def write_index_sheet(
    wb: Workbook,
    validation: pd.DataFrame,
    allocation: pd.DataFrame,
    report_tabs: list[str],
) -> None:
    ws = wb.create_sheet("Index", 1)
    ws.sheet_view.showGridLines = False
    report_tab_set = set(report_tabs)
    summary = (
        validation.loc[validation["tab"].isin(report_tab_set)]
        .groupby("tab", dropna=False)
        .agg(
            Source_Route_BSA=("route_bsa", "sum"),
            Allocated_BSA=("allocated_bsa", "sum"),
            BSA_Diff=("bsa_diff", "sum"),
            LST_TEU=("lst", "sum"),
            WOS3_FST_TEU=("w3", "sum"),
            WOS3_LST_TEU=("w3_norm_lst", "sum"),
            High_Profit_WOS3_FST_TEU=("hi_w3", "sum"),
        )
        .reset_index()
        .rename(columns={"tab": "Tab"})
    )
    no_basis = (
        allocation.loc[allocation["allocation_level"].eq(NO_BASIS_LEVEL)]
        .groupby("tab", dropna=False)
        .agg(No_Basis_Rows=("allocated_bsa", "size"), No_Basis_BSA=("allocated_bsa", "sum"))
        .reset_index()
        .rename(columns={"tab": "Tab"})
    )
    summary = summary.merge(no_basis, on="Tab", how="left")
    summary[["No_Basis_Rows", "No_Basis_BSA"]] = summary[["No_Basis_Rows", "No_Basis_BSA"]].fillna(0)
    summary["Has_Report_Tab"] = summary["Tab"].isin(report_tab_set).map({True: "Y", False: "N"})
    summary["WOS3_BSA"] = summary["WOS3_FST_TEU"] / summary["Allocated_BSA"].replace(0, pd.NA)
    summary["WOS3_Shipment_Rate"] = summary["WOS3_LST_TEU"] / summary["WOS3_FST_TEU"].replace(0, pd.NA)
    summary["High_Profit_WOS3_Share"] = summary["High_Profit_WOS3_FST_TEU"] / summary["WOS3_FST_TEU"].replace(0, pd.NA)
    summary = summary[
        [
            "Tab",
            "Has_Report_Tab",
            "Source_Route_BSA",
            "Allocated_BSA",
            "BSA_Diff",
            "LST_TEU",
            "WOS3_FST_TEU",
            "WOS3_BSA",
            "WOS3_LST_TEU",
            "WOS3_Shipment_Rate",
            "High_Profit_WOS3_FST_TEU",
            "High_Profit_WOS3_Share",
            "No_Basis_Rows",
            "No_Basis_BSA",
        ]
    ].sort_values(["Has_Report_Tab", "Source_Route_BSA", "LST_TEU", "Tab"], ascending=[False, False, False, True])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    no_basis_header_fill = PatternFill("solid", fgColor="F4B183")
    no_basis_value_fill = PatternFill("solid", fgColor="FCE4D6")
    no_basis_headers = {"No_Basis_Rows", "No_Basis_BSA"}
    ws.cell(1, 1, "3W Booking Action Index")
    ws.cell(1, 1).font = Font(bold=True, size=14)
    for col, header in enumerate(summary.columns, start=1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        if header in no_basis_headers:
            cell.fill = no_basis_header_fill
            cell.comment = Comment(NO_BASIS_NOTE, "Codex")
    for r_idx, row in enumerate(summary.itertuples(index=False), start=4):
        tab_name = row.Tab
        for c_idx, value in enumerate(row, start=1):
            value = excel_value(value)
            cell = ws.cell(r_idx, c_idx, value)
            if c_idx == 1 and tab_name in report_tab_set:
                ref = str(tab_name).replace("'", "''")
                label = str(tab_name).replace('"', '""')
                cell.value = f'=HYPERLINK("#\'{ref}\'!A1","{label}")'
                cell.style = "Hyperlink"
            if c_idx in {8, 10, 12}:
                cell.number_format = "0.0%"
            elif c_idx >= 3:
                cell.number_format = "#,##0.0"
            if summary.columns[c_idx - 1] in no_basis_headers:
                cell.fill = no_basis_value_fill
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(summary.shape[1])}{summary.shape[0] + 3}"
    widths = [16, 14, 16, 16, 12, 14, 14, 12, 14, 16, 22, 16, 14, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_validation(
    bsa: pd.DataFrame,
    allocation: pd.DataFrame,
    lst: pd.DataFrame,
    w3: pd.DataFrame,
    w3_norm_lst: pd.DataFrame,
    hi_w3: pd.DataFrame,
) -> pd.DataFrame:
    bsa_source = bsa.groupby(["tab", "yyyymm"], dropna=False)["route_bsa"].sum().reset_index()
    bsa_alloc = allocation.groupby(["tab", "yyyymm"], dropna=False)["allocated_bsa"].sum().reset_index()
    out = bsa_source.merge(bsa_alloc, on=["tab", "yyyymm"], how="outer").fillna(0)
    out["bsa_diff"] = out["allocated_bsa"] - out["route_bsa"]

    for name, frame in [("lst", lst), ("w3", w3), ("w3_norm_lst", w3_norm_lst), ("hi_w3", hi_w3)]:
        metric = frame.groupby(["tab", "yyyymm"], dropna=False)["value"].sum().reset_index()
        out = out.merge(metric.rename(columns={"value": name}), on=["tab", "yyyymm"], how="outer")
    for col in ["lst", "w3", "w3_norm_lst", "hi_w3"]:
        out[col] = out[col].fillna(0.0)
    out["w3_bsa"] = out["w3"] / out["allocated_bsa"].replace(0, pd.NA)
    out["w3_ship_rate"] = out["w3_norm_lst"] / out["w3"].replace(0, pd.NA)
    out["hi_w3_share"] = out["hi_w3"] / out["w3"].replace(0, pd.NA)
    return out.sort_values(["tab", "yyyymm"]).reset_index(drop=True)


def write_readme(wb: Workbook, xlsx_name: str, tab_count: int) -> None:
    ws = wb.active
    ws.title = "README"
    ws.sheet_view.showGridLines = False
    rows = [
        ("Workbook", xlsx_name),
        ("Generated at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Source booking cache", f"_cache_2025.parquet + _cache_{CURRENT_DATASET_ID}.parquet"),
        ("Source BSA", f"BSA_raw_monthly3W_{CURRENT_DATASET_ID}.csv"),
        ("Report months", "2025-01 to 2026-03"),
        ("Report tabs", tab_count),
        ("Index", "Tab-level BSA, LST, WOS-3, high-profit share, no-basis checks"),
        ("Tab rule", "CN by POR port with SHK+DCB combined; VN as SGN+CMP and HPH; others by origin country"),
        ("LST performance", "LST_Status=Normal, LST_TEU, dashboard 445 YYYYMM"),
        ("BSA allocation", "Monthly route BSA allocated only by exact POR_port + DLY_port 2025 Normal LST_TEU salesperson share"),
        (
            "Allocation fallback",
            "No CN/origin-wide fallback; routes without exact 2025 basis stay in (no 2025 basis)",
        ),
        ("3W booking", "Lead_time (BKG_Sche)=WOS-3, FST_TEU"),
        ("3W shipment", "Lead_time (BKG_Sche)=WOS-3 and LST_Status=Normal, LST_TEU"),
        ("High-profit 3W booking", "Dashboard profit type 고/저=고수익 within WOS-3 FST_TEU"),
        ("Ratio totals", "Quarter/year ratios use summed numerator divided by summed denominator"),
    ]
    ws.cell(1, 1, "3W Booking Action BSA Salesperson Sheet")
    ws.cell(1, 1).font = Font(bold=True, size=14)
    no_basis_fill = PatternFill("solid", fgColor="FCE4D6")
    for r_idx, (key, value) in enumerate(rows, start=3):
        key_cell = ws.cell(r_idx, 1, key)
        value_cell = ws.cell(r_idx, 2, value)
        key_cell.font = Font(bold=True)
        if key == "Allocation fallback":
            key_cell.fill = no_basis_fill
            value_cell.fill = no_basis_fill
            value_cell.comment = Comment(NO_BASIS_NOTE, "Codex")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 110


def upload_as_google_sheet(xlsx_path: Path) -> dict[str, str]:
    try:
        from google.auth.transport.requests import Request
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaFileUpload
    except Exception as exc:
        return {"error": f"Google API library unavailable: {exc}"}

    creds_dir = ROOT.parent / ".gdrive-mcp"
    credentials_path = creds_dir / "credentials.json"
    token_path = creds_dir / "token.json"
    if not credentials_path.exists() or not token_path.exists():
        return {"error": "Google Drive credentials were not found"}

    installed = json.loads(credentials_path.read_text(encoding="utf-8-sig"))["installed"]
    token = json.loads(token_path.read_text(encoding="utf-8-sig"))
    creds = Credentials(
        token=token.get("access_token"),
        refresh_token=token["refresh_token"],
        token_uri=installed["token_uri"],
        client_id=installed["client_id"],
        client_secret=installed["client_secret"],
    )
    creds.refresh(Request())
    service = build("drive", "v3", credentials=creds)
    media = MediaFileUpload(
        str(xlsx_path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True,
    )
    folder_id = os.environ.get("GDRIVE_FOLDER_ID", "1JIxg6Y-_gRfI1HueXZ1Q9j4-Z5bxvNgv")
    body = {
        "name": "3W Booking Action BSA Salesperson OBT 2025-2026Q1",
        "mimeType": "application/vnd.google-apps.spreadsheet",
        "parents": [folder_id],
    }
    created = service.files().create(
        body=body,
        media_body=media,
        fields="id,name,webViewLink",
        supportsAllDrives=True,
    ).execute()
    try:
        apply_google_sheet_index_links(creds, created["id"])
        created["index_links"] = "gid formulas applied"
    except Exception as exc:
        created["index_links_error"] = str(exc)
    return created


def apply_google_sheet_index_links(creds: object, spreadsheet_id: str) -> None:
    from googleapiclient.discovery import build

    service = build("sheets", "v4", credentials=creds)
    meta = service.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets(properties(sheetId,title))",
    ).execute()
    title_to_gid = {
        sheet["properties"]["title"]: sheet["properties"]["sheetId"]
        for sheet in meta.get("sheets", [])
    }
    values = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range="Index!A4:A",
        valueRenderOption="FORMATTED_VALUE",
    ).execute().get("values", [])
    formulas: list[list[str]] = []
    for row in values:
        label = row[0] if row else ""
        if label in title_to_gid:
            safe_label = label.replace('"', '""')
            formulas.append([f'=HYPERLINK("#gid={title_to_gid[label]}","{safe_label}")'])
        elif label:
            formulas.append([label])
        else:
            break
    if formulas:
        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"Index!A4:A{len(formulas) + 3}",
            valueInputOption="USER_ENTERED",
            body={"values": formulas},
        ).execute()


def main() -> None:
    OUT_DIR.mkdir(exist_ok=True)
    booking = load_booking()
    bsa = load_bsa()
    basis_detail, lookups = build_basis(booking)
    allocation = allocate_bsa(bsa, lookups)

    lst = metric_sum(booking, booking["status"].eq("Normal") & booking["lst"].gt(0), "lst")
    w3_mask = booking["lead_time"].eq("WOS-3") & booking["fst"].gt(0)
    w3 = metric_sum(booking, w3_mask, "fst")
    w3_norm_source = booking.assign(
        _w3_norm_lst=booking["lst"].where(w3_mask & booking["status"].eq("Normal"), 0.0)
    )
    w3_norm_lst = metric_sum(w3_norm_source, w3_norm_source["_w3_norm_lst"].gt(0), "_w3_norm_lst")
    hi_w3 = metric_sum(w3_source := booking.assign(_hi_w3=booking["fst"].where(w3_mask & booking["profit_type"].eq("고수익"), 0.0)), w3_source["_hi_w3"].gt(0), "_hi_w3")
    bsa_sales = sales_metric_from_allocation(allocation)

    tab_names = sorted(set(bsa["tab"].unique()) | set(lst["tab"].unique()) | set(w3["tab"].unique()))
    bsa_tabs = set(bsa["tab"].unique())
    tab_names = [tab for tab in tab_names if tab in bsa_tabs]

    metrics: dict[str, pd.DataFrame | tuple[pd.DataFrame, pd.DataFrame]] = {
        "lst": lst,
        "bsa": bsa_sales,
        "w3": w3,
        "w3_norm_lst": w3_norm_lst,
        "w3_ship_rate": (w3_norm_lst, w3),
        "w3_bsa": (w3, bsa_sales),
        "hi_w3": hi_w3,
        "hi_share": (hi_w3, w3),
    }
    validation = build_validation(bsa, allocation, lst, w3, w3_norm_lst, hi_w3)
    validation = validation.loc[validation["tab"].isin(tab_names)].copy()

    xlsx_path = OUT_DIR / "3W_Booking_Action_BSA_Salesperson_OBT_2025_2026Q1.xlsx"
    wb = Workbook()
    write_readme(wb, xlsx_path.name, len(tab_names))
    write_index_sheet(wb, validation, allocation, tab_names)
    existing = {"README", "Index"}
    for tab in tab_names:
        rows = build_rows_for_tab(tab, [lst, bsa_sales, w3, w3_norm_lst, hi_w3], lst)
        write_report_sheet(wb, safe_sheet_title(tab, existing), rows, metrics)

    allocation_out = allocation.rename(
        columns={
            "yyyymm": "YYYYMM",
            "tab": "Tab",
            "team": "Team",
            "origin": "POR_Country",
            "ori_port": "POR_Port",
            "dest": "DLY_Country",
            "dst_port": "DLY_Port",
            "sales": "Salesman",
            "basis_lST_TEU": "2025_Basis_LST_TEU",
            "basis_total_lST_TEU": "2025_Basis_Total_LST_TEU",
            "allocation_share": "Allocation_Share",
            "route_bsa": "Route_BSA",
            "allocated_bsa": "Allocated_BSA",
            "allocation_level": "Allocation_Level",
            "allocation_key": "Allocation_Key",
        }
    )
    basis_out = basis_detail.rename(
        columns={
            "tab": "Tab",
            "team": "Team",
            "origin": "POR_Country",
            "ori_port": "POR_Port",
            "dest": "DLY_Country",
            "dst_port": "DLY_Port",
            "shipper_code": "Shipper_Code",
            "shipper_name": "Shipper_Name",
            "sales": "Salesman",
        }
    )
    validation_out = validation.rename(
        columns={
            "tab": "Tab",
            "yyyymm": "YYYYMM",
            "route_bsa": "Source_Route_BSA",
            "allocated_bsa": "Allocated_BSA",
            "bsa_diff": "BSA_Diff",
            "lst": "LST_TEU",
            "w3": "WOS3_FST_TEU",
            "w3_norm_lst": "WOS3_LST_TEU",
            "hi_w3": "High_Profit_WOS3_FST_TEU",
            "w3_bsa": "WOS3_BSA",
            "w3_ship_rate": "WOS3_Shipment_Rate",
            "hi_w3_share": "High_Profit_WOS3_Share",
        }
    )

    write_dataframe_sheet(wb, "BSA_Backdata", allocation_out)
    write_dataframe_sheet(wb, "2025_Basis", basis_out)
    write_dataframe_sheet(wb, "Validation", validation_out)
    wb.save(xlsx_path)

    upload_result = upload_as_google_sheet(xlsx_path)
    report = {
        "xlsx_path": str(xlsx_path),
        "google_sheet": upload_result,
        "tabs": tab_names,
        "tab_count": len(tab_names),
        "allocation_rows": int(len(allocation_out)),
        "basis_rows": int(len(basis_out)),
        "validation": {
            "source_bsa": float(validation["route_bsa"].sum()),
            "allocated_bsa": float(validation["allocated_bsa"].sum()),
            "bsa_diff": float(validation["allocated_bsa"].sum() - validation["route_bsa"].sum()),
            "wos3_fst_teu": float(validation["w3"].sum()),
            "wos3_lst_teu": float(validation["w3_norm_lst"].sum()),
            "wos3_shipment_rate": (
                float(validation["w3_norm_lst"].sum()) / float(validation["w3"].sum())
                if float(validation["w3"].sum()) else None
            ),
            "no_2025_basis_route_rows": int((allocation["allocation_level"] == NO_BASIS_LEVEL).sum()),
            "no_2025_basis_bsa": float(
                allocation.loc[allocation["allocation_level"].eq(NO_BASIS_LEVEL), "allocated_bsa"].sum()
            ),
        },
    }
    report_path = OUT_DIR / "3W_Booking_Action_BSA_Salesperson_2025_2026Q1_report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"xlsx={xlsx_path}")
    print(f"report={report_path}")
    print(f"google_sheet={upload_result}")
    print(f"validation={report['validation']}")


if __name__ == "__main__":
    main()
