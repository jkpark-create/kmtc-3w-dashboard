from __future__ import annotations

import argparse
import importlib.util
import io
import json
import re
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "output"

SALES_OWNER_SOURCE_ID = "1T1wUR8qmvbGL_WLeD3tpQnsuOeQ7vHwQoNFWEVKEiyM"
REPORT_SOURCE_ID = "19d_lnB7Qt6H-UJE7ECyDo-i385tjMjAX8aATk0R0N54"
TARGET_SPREADSHEET_ID = "1YxZkwvoMaQXIEw07qUDZtCPDFZBf8GOZyr5knkxnLxo"
SALES_OWNER_INPUT_SHEET = "Sales_Owner_Input"
MISSING_SALES = "(\ubbf8\uc9c0\uc815)"


# Target workbook tab -> source organization-chart sheet / column / row span.
# Only sales/marketing columns are included; other departments are intentionally excluded.
SALES_OWNER_RANGES: dict[str, list[tuple[str, str, int, int]]] = {
    "CN_SHA": [("SHA", "H", 19, 24)],
    "CN_NKG": [("YZR", "F", 17, 19)],
    "CN_NBO": [("NBO", "G", 16, 18)],
    "CN_TAO": [("TAO", "G", 17, 20)],
    "CN_LYG": [("TAO", "J", 18, 18)],
    "CN_XGG": [("XGG", "G", 16, 18)],
    "CN_DLC": [("DLC", "F", 16, 18)],
    "CN_NNS": [("CAN", "G", 18, 20)],
    "CN_SHK_DCB": [("SZP", "G", 17, 20)],
    "CN_XMN": [("XMN", "F", 18, 20)],
    "HK": [("HK", "J", 22, 26)],
    "SG": [("SG", "G", 18, 21)],
    "TH": [("TH", "H", 21, 26)],
    "VN_SGN_CMP": [("VN", "E", 19, 22)],
    "VN_HPH": [("VN", "V", 19, 21)],
    "PH": [("PH", "H", 17, 20)],
    "IN": [("IN", "I", 22, 26), ("IN", "L", 21, 27), ("IN", "P", 22, 27), ("IN", "T", 21, 27)],
    "AE": [("AE", "F", 16, 20)],
    "JKT": [("ID", "F", 18, 24)],
    "SUB": [("ID", "Q", 20, 24)],
    "ID_out": [("ID", "G", 20, 24)],
    "PKG+PKW": [("MY", "N", 21, 23)],
    "PGU": [("MY", "R", 23, 23)],
    "PEN": [("MY", "W", 23, 23)],
}


MANUAL_ALIASES = {
    "JENNYDOU": "JENNY",
    "CRUISESONG": "CRUISE",
    "TERRYWANG": "TERRY",
    "COLINYANG": "COLIN",
    "BONOYUAN": "BONO",
    "CHLOEZHOU": "CHLOE",
    "SCOTTGAO": "SCOTT",
    "BILLGUAN": "BILL",
    "ROSSFU": "ROSS",
    "MARTINCHANG": "MARTIN1",
    "JASONWU": "JCWU",
    "JOELI": "CANID",
    "CHAMHM": "CHANHM",
    "PETERKONG": "KONG",
    "DARRENSIM": "DARREN",
    "ALEXLIN": "ALEX",
    "KELVINTAY": "KVIN",
    "ALVINANG": "ALVINA",
    "AKBAR": "AKBARA",
    "KEVIN": "KEVINYESAY",
    "RINATAN": "RINA",
    "EDWINKHU": "EDWIN",
    "JESSQAWOO": "JESSIQA",
    "TEHTL": "TLTEH",
}


EXCLUDE_KEYS = {
    "SALES",
    "MARKETING",
    "SALESMARKETING",
    "SALESDEPARTMENT",
    "SALESDIVISION",
    "MKTSLS",
    "SALESCS",
    "SALESC",
    "CS",
    "OBSALES",
    "IBSALES",
    "BSALES",
    "TOTAL",
    "NUMBER",
    "CONFIDENTIAL",
    "OUTPORT",
}


@dataclass
class OwnerEntry:
    target_tab: str
    source_sheet: str
    source_cell: str
    input_name: str
    sales_key: str
    resolved_sales: str
    match_status: str


def load_target_builder():
    path = ROOT / "scripts" / "create_per_origin_target_workbook.py"
    spec = importlib.util.spec_from_file_location("target_builder", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def get_creds() -> Credentials:
    creds_dir = ROOT.parent / ".gdrive-mcp"
    credentials_path = creds_dir / "credentials.json"
    token_path = creds_dir / "token.json"
    installed = json.loads(credentials_path.read_text(encoding="utf-8-sig"))["installed"]
    token = json.loads(token_path.read_text(encoding="utf-8-sig"))
    creds = Credentials(
        token=token.get("access_token"),
        refresh_token=token["refresh_token"],
        token_uri=installed["token_uri"],
        client_id=installed["client_id"],
        client_secret=installed["client_secret"],
    )
    creds.refresh(Request())
    return creds


def with_backoff(call: Any) -> Any:
    for attempt in range(7):
        try:
            return call.execute()
        except Exception as exc:  # noqa: BLE001
            if "429" not in str(exc) or attempt == 6:
                raise
            time.sleep((2**attempt) * 5)
    raise RuntimeError("unreachable")


def export_sheet(drive: Any, file_id: str, filename: str) -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = OUT_DIR / filename
    request = drive.files().export_media(
        fileId=file_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    with io.FileIO(out, "wb") as fh:
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
    return out


def norm_key(value: object) -> str:
    return re.sub(r"[^A-Z0-9가-힣]", "", str(value or "").upper())


def colnum(letter: str) -> int:
    out = 0
    for ch in letter:
        out = out * 26 + ord(ch.upper()) - 64
    return out


def split_names(value: object) -> list[str]:
    if value is None:
        return []
    text = str(value).replace("\xa0", " ").replace("*", "").strip()
    text = re.sub(
        r"\([^)]*(?:DRIV|DRIVER|IT|HR|IB\.|OB\.|additional post|Messenger|Maid)[^)]*\)",
        "",
        text,
        flags=re.I,
    )
    parts = re.split(r"\s*/\s*|\s*,\s*|\n+|\s{2,}", text)
    names: list[str] = []
    for part in parts:
        part = re.sub(r"\([^)]*\)", "", part).strip()
        part = re.sub(r"\s+", " ", part)
        key = norm_key(part)
        if not key or key.isdigit() or key in EXCLUDE_KEYS or len(key) <= 1:
            continue
        names.append(part)
    return names


def workbook_values(path: Path, *, read_only: bool = True) -> dict[str, list[list[Any]]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=read_only)
    out: dict[str, list[list[Any]]] = {}
    for ws in wb.worksheets:
        out[ws.title] = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
    return out


def read_existing_target_input(target_path: Path, mod: Any) -> dict[str, tuple[float, float, float, str]]:
    wb = openpyxl.load_workbook(target_path, data_only=True, read_only=True)
    if mod.INPUT_SHEET not in wb.sheetnames:
        return {}
    ws = wb[mod.INPUT_SHEET]
    existing: dict[str, tuple[float, float, float, str]] = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        tab = str(row[0]).strip() if row and row[0] is not None else ""
        if not tab:
            continue
        booking = mod.clean_number(row[1]) or 0.0
        lifting = mod.clean_number(row[3]) or 0.0
        hp = mod.clean_number(row[5]) or 0.0
        memo = "" if len(row) < 8 or row[7] is None else str(row[7])
        existing[tab] = (booking, lifting, hp, memo)
    return existing


def load_owner_entries(owner_path: Path, report_tabs: dict[str, list[list[Any]]], mod: Any) -> dict[str, list[OwnerEntry]]:
    wb = openpyxl.load_workbook(owner_path, data_only=True, read_only=False)
    parsed: dict[str, list[OwnerEntry]] = {}
    for target_tab, ranges in SALES_OWNER_RANGES.items():
        seen: set[str] = set()
        raw_rows: list[dict[str, str]] = []
        for sheet, col, start, end in ranges:
            ws = wb[sheet]
            c = colnum(col)
            for row in range(start, end + 1):
                for name in split_names(ws.cell(row, c).value):
                    key = norm_key(name)
                    if key in seen:
                        continue
                    seen.add(key)
                    raw_rows.append(
                        {
                            "target_tab": target_tab,
                            "source_sheet": sheet,
                            "source_cell": ws.cell(row, c).coordinate,
                            "input_name": name,
                            "sales_key": key,
                        }
                    )
        blocks = mod.parse_report_tab(report_tabs.get(target_tab, []))
        available = sorted({name for block in blocks.values() for name in block if name not in {"TOTAL", mod.NO_BASIS_LABEL}})
        available_by_key = {norm_key(name): name for name in available}
        used: set[str] = set()
        entries: list[OwnerEntry] = []
        for row in raw_rows:
            resolved, status = resolve_sales(row["sales_key"], available_by_key, used)
            used.add(resolved)
            entries.append(OwnerEntry(resolved_sales=resolved, match_status=status, **row))
        parsed[target_tab] = entries
    return parsed


def resolve_sales(key: str, available_by_key: dict[str, str], used: set[str]) -> tuple[str, str]:
    if key in available_by_key:
        return available_by_key[key], "exact"
    alias = MANUAL_ALIASES.get(key)
    if alias and alias in available_by_key:
        return available_by_key[alias], "alias"

    candidates: list[tuple[str, str]] = []
    for avail_key, name in available_by_key.items():
        if name in used:
            continue
        if len(avail_key) >= 4 and (key.startswith(avail_key) or key.endswith(avail_key)):
            candidates.append((avail_key, name))
        elif len(key) >= 4 and avail_key.startswith(key):
            candidates.append((avail_key, name))
    if len(candidates) == 1:
        return candidates[0][1], "prefix"
    return key, "unmatched"


def raw_for_sales(mod: Any, sales: str, blocks: dict[str, dict[str, list[Any]]]) -> dict[str, float]:
    if not sales:
        return {k: 0.0 for k in mod.RAW_KEYS}
    return mod.build_raw_metrics(sales, blocks)


def collect_rows_for_origin(
    mod: Any,
    origin: str,
    blocks: dict[str, dict[str, list[Any]]],
    account_counts: dict[tuple[str, str], tuple[int, int, float | None]],
    owner_entries: dict[str, list[OwnerEntry]],
) -> list[dict[str, Any]]:
    entries = owner_entries.get(origin, [])
    if not entries:
        return mod.collect_rows_for_origin(origin, blocks, account_counts)

    raw_by_sales = {entry.resolved_sales: raw_for_sales(mod, entry.resolved_sales, blocks) for entry in entries}
    team_raw = {k: 0.0 for k in mod.RAW_KEYS}
    for raw in raw_by_sales.values():
        for key in mod.RAW_KEYS:
            team_raw[key] += raw[key]

    rows = [mod.build_display_row(origin, "Team Total", team_raw, team_raw, is_total=True)]
    for entry in entries:
        rows.append(mod.build_display_row(origin, entry.resolved_sales, raw_by_sales[entry.resolved_sales], team_raw, is_total=False))
    return rows


def compute_suggestions(
    mod: Any,
    report_tabs: dict[str, list[list[Any]]],
    account_counts: dict[tuple[str, str], tuple[int, int, float | None]],
    origins: list[str],
    owner_entries: dict[str, list[OwnerEntry]],
) -> dict[str, dict[str, Any]]:
    suggestions: dict[str, dict[str, Any]] = {}
    for origin in origins:
        blocks = mod.parse_report_tab(report_tabs.get(origin, []))
        if not {"lst", "bsa", "w3", "w3_norm_lst", "hi_w3"}.issubset(blocks):
            continue
        rows = collect_rows_for_origin(mod, origin, blocks, account_counts, owner_entries)
        team = rows[0]
        metrics = {
            "booking": {
                "base": team["booking_base_2025"],
                "perf": team["booking_q1_perform"],
                "is_hp": False,
            },
            "lifting": {
                "base": team["lifting_base_2025"],
                "perf": team["lifting_q1_perform"],
                "is_hp": False,
            },
            "hp": {
                "base": team["high_profit_base_2025"],
                "perf": team["high_profit_q1_perform"],
                "is_hp": True,
            },
        }
        out_metrics: dict[str, Any] = {}
        for name, metric in metrics.items():
            pp, why = mod.suggest_pp(metric["base"], metric["perf"], is_hp=metric["is_hp"])
            out_metrics[name] = {"base": metric["base"], "perf": metric["perf"], "pp": pp, "why": why}
        suggestions[origin] = {
            "booking_pp": out_metrics["booking"]["pp"],
            "lifting_pp": out_metrics["lifting"]["pp"],
            "hp_pp": out_metrics["hp"]["pp"],
            "metrics": out_metrics,
        }
    return suggestions


def build_summary_values(
    mod: Any,
    report_tabs: dict[str, list[list[Any]]],
    account_counts: dict[tuple[str, str], tuple[int, int, float | None]],
    origins: list[str],
    owner_entries: dict[str, list[OwnerEntry]],
) -> list[list[Any]]:
    values: list[list[Any]] = [["2026 OBT Sales Target - All origins"]]
    values.extend(mod.header_block(include_tab_col=True))
    target_cols = ("E", "H", "K", "N", "Q", "T")
    perform_for_diff = ("F", "I", "L", "O", "R", "U")
    base_cols = {"booking": "Z", "lifting": "AA", "hp": "AB"}
    for origin in origins:
        blocks = mod.parse_report_tab(report_tabs.get(origin, []))
        if not {"lst", "bsa", "w3", "w3_norm_lst", "hi_w3"}.issubset(blocks):
            continue
        rows = collect_rows_for_origin(mod, origin, blocks, account_counts, owner_entries)
        for item in rows:
            sheet_row = len(values) + 1
            counts = account_counts.get((origin, item["sales"]), (None, None, None))
            values.append(
                mod.make_data_row(
                    item=item,
                    sheet_row=sheet_row,
                    counts=counts,
                    base_cols=base_cols,
                    origin_literal=None,
                    target_cols=target_cols,
                    perform_cols=perform_for_diff,
                    leading=[origin],
                )
            )
    return values


def build_owner_input_values(owner_entries: dict[str, list[OwnerEntry]], origins: list[str]) -> list[list[Any]]:
    values: list[list[Any]] = [
        ["Sales owner input"],
        [f"Generated at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} from {SALES_OWNER_SOURCE_ID}"],
        ["Active", "Target_Tab", "Salesman", "Input_Name", "Metric_Salesman", "Match_Status", "Source_Sheet", "Source_Cell", "Memo"],
    ]
    for origin in origins:
        for entry in owner_entries.get(origin, []):
            values.append(
                [
                    "Y",
                    entry.target_tab,
                    entry.resolved_sales,
                    entry.input_name,
                    entry.resolved_sales,
                    entry.match_status,
                    entry.source_sheet,
                    entry.source_cell,
                    "",
                ]
            )
        if origin not in owner_entries:
            values.append(["", origin, "", "", "", "no_source_list", "", "", "Existing report-derived rows retained."])
    return values


def ensure_support_sheet(service: Any, spreadsheet_id: str) -> dict[str, int]:
    meta = with_backoff(
        service.spreadsheets().get(
            spreadsheetId=spreadsheet_id,
            fields="sheets(properties(sheetId,title,index))",
        )
    )
    sheet_ids = {s["properties"]["title"]: s["properties"]["sheetId"] for s in meta.get("sheets", [])}
    if SALES_OWNER_INPUT_SHEET not in sheet_ids:
        with_backoff(
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={
                    "requests": [
                        {
                            "addSheet": {
                                "properties": {
                                    "title": SALES_OWNER_INPUT_SHEET,
                                    "index": 2,
                                    "gridProperties": {"rowCount": 500, "columnCount": 12},
                                }
                            }
                        }
                    ]
                },
            )
        )
        meta = with_backoff(
            service.spreadsheets().get(
                spreadsheetId=spreadsheet_id,
                fields="sheets(properties(sheetId,title,index))",
            )
        )
        sheet_ids = {s["properties"]["title"]: s["properties"]["sheetId"] for s in meta.get("sheets", [])}
    return sheet_ids


def ordered_target_origins(service: Any, spreadsheet_id: str, mod: Any, report_tabs: dict[str, list[list[Any]]]) -> list[str]:
    meta = with_backoff(
        service.spreadsheets().get(
            spreadsheetId=spreadsheet_id,
            fields="sheets(properties(title,index))",
        )
    )
    support = {mod.README_SHEET, mod.INPUT_SHEET, mod.SUMMARY_SHEET, SALES_OWNER_INPUT_SHEET}
    sheets = sorted(meta.get("sheets", []), key=lambda s: s["properties"].get("index", 0))
    origins = []
    for sheet in sheets:
        title = sheet["properties"]["title"]
        if title in support:
            continue
        if title in report_tabs:
            origins.append(title)
    return origins


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Build payloads and audit files without writing to Google Sheets.")
    args = parser.parse_args()

    mod = load_target_builder()
    creds = get_creds()
    drive = build("drive", "v3", credentials=creds)
    sheets = build("sheets", "v4", credentials=creds)

    owner_path = export_sheet(drive, SALES_OWNER_SOURCE_ID, "sales_owner_source_export_20260519.xlsx")
    report_path = export_sheet(drive, REPORT_SOURCE_ID, "salesperson_bsa_source_report_export_20260519.xlsx")
    target_path = export_sheet(drive, TARGET_SPREADSHEET_ID, "target_workbook_before_sales_owner_update_20260519.xlsx")

    report_values_all = workbook_values(report_path)
    report_tabs = {
        title: values
        for title, values in report_values_all.items()
        if title not in mod.SUPPORT_TABS and mod.is_report_values(values)
    }
    account_counts = mod.build_account_counts()
    existing_input = read_existing_target_input(target_path, mod)
    owner_entries = load_owner_entries(owner_path, report_tabs, mod)
    origins = ordered_target_origins(sheets, TARGET_SPREADSHEET_ID, mod, report_tabs)
    suggestions = compute_suggestions(mod, report_tabs, account_counts, origins, owner_entries)

    summary_values = build_summary_values(mod, report_tabs, account_counts, origins, owner_entries)
    payloads: list[tuple[str, list[list[Any]]]] = [
        (mod.README_SHEET, mod.build_readme_values()),
        (mod.INPUT_SHEET, mod.build_input_values(origins, existing_input, suggestions)),
        (SALES_OWNER_INPUT_SHEET, build_owner_input_values(owner_entries, origins)),
        (mod.SUMMARY_SHEET, summary_values),
    ]

    origin_rowcounts: dict[str, int] = {}
    for origin in origins:
        blocks = mod.parse_report_tab(report_tabs[origin])
        rows = collect_rows_for_origin(mod, origin, blocks, account_counts, owner_entries)
        values = mod.build_origin_sheet_values(origin, rows, account_counts)
        payloads.append((origin, values))
        origin_rowcounts[origin] = len(values)

    audit_rows = []
    for origin in origins:
        for entry in owner_entries.get(origin, []):
            audit_rows.append(
                {
                    "Target_Tab": entry.target_tab,
                    "Salesman": entry.resolved_sales,
                    "Input_Name": entry.input_name,
                    "Match_Status": entry.match_status,
                    "Source_Sheet": entry.source_sheet,
                    "Source_Cell": entry.source_cell,
                }
            )
    audit_path = OUT_DIR / "sales_owner_input_audit_20260519.csv"
    import pandas as pd

    pd.DataFrame(audit_rows).to_csv(audit_path, index=False, encoding="utf-8-sig")

    report = {
        "target_spreadsheet_id": TARGET_SPREADSHEET_ID,
        "sales_owner_source_id": SALES_OWNER_SOURCE_ID,
        "report_source_id": REPORT_SOURCE_ID,
        "origins": len(origins),
        "origin_rowcounts": origin_rowcounts,
        "owner_rows": len(audit_rows),
        "matched_owner_rows": sum(1 for row in audit_rows if row["Match_Status"] != "unmatched"),
        "unmatched_owner_rows": sum(1 for row in audit_rows if row["Match_Status"] == "unmatched"),
        "audit_csv": str(audit_path),
        "dry_run": args.dry_run,
    }
    report_path_json = OUT_DIR / "sales_owner_target_update_20260519.json"
    report_path_json.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))

    if args.dry_run:
        return

    ensure_support_sheet(sheets, TARGET_SPREADSHEET_ID)
    mod.clear_tab_values(
        sheets,
        TARGET_SPREADSHEET_ID,
        [mod.README_SHEET, mod.INPUT_SHEET, SALES_OWNER_INPUT_SHEET, mod.SUMMARY_SHEET, *origins],
    )
    mod.batch_write_values(sheets, TARGET_SPREADSHEET_ID, payloads)


if __name__ == "__main__":
    main()
