from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "output"
DATASET_ID = "20260508"
MONTHS = ["202603", "202604", "202605"]
MONTH_LABELS = {"202603": "3월", "202604": "4월", "202605": "5월"}
ORIGINS = ["TH", "VN"]
TEAM = "OBT"
DEST_ROWS = {
    "TH": ["CN", "VN", "JP", "ID", "IN"],
    "VN": ["CN", "JP", "ID", "TH", "IN"],
}


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [re.sub(r"[^\x00-\x7F]+$", "", c).strip() for c in out.columns]
    return out


def to_number(series: pd.Series) -> pd.Series:
    return pd.to_numeric(
        series.fillna("").astype(str).str.replace(",", "", regex=False).str.strip(),
        errors="coerce",
    ).fillna(0.0)


def classify_team(origin: str, dest: str) -> str:
    origin = str(origin).strip()
    dest = str(dest).strip()
    if origin not in ("KR", "JP") and dest != "KR":
        return "OBT"
    if origin == "KR" and dest != "JP":
        return "EST"
    if origin != "JP" and dest == "KR":
        return "IST"
    return "JBT"


def norm_text(value: object, fallback: str = "") -> str:
    text = "" if pd.isna(value) else str(value).strip()
    if text.lower() in ("nan", "none", "nat"):
        return fallback
    return text or fallback


def pct_value(num: float, den: float) -> float | None:
    if den and den != 0:
        return float(num) / float(den)
    return None


def pct_text(value: float | None) -> str:
    if value is None or pd.isna(value):
        return "-"
    return f"{round(value * 100):.0f}%"


def trend_text(values: list[float | None]) -> str:
    numeric = [v for v in values if v is not None and not pd.isna(v)]
    if not numeric:
        return "-"
    lo = min(numeric)
    hi = max(numeric)
    blocks = "▁▂▃▄▅▆▇█"
    out = []
    for value in values:
        if value is None or pd.isna(value):
            out.append("·")
        elif hi == lo:
            out.append("▄")
        else:
            idx = round((value - lo) / (hi - lo) * (len(blocks) - 1))
            out.append(blocks[int(idx)])
    return "".join(out)


def load_current_summary() -> dict:
    path = OUT_DIR / f"dashboard_summary_{DATASET_ID}.json"
    if not path.exists():
        raise FileNotFoundError(f"Current summary not found: {path}")
    with path.open(encoding="utf-8") as fh:
        return json.load(fh)


def load_2025_basis() -> pd.DataFrame:
    df2_cols = [
        "BKG_NO",
        "POR_Country",
        "POR_PORT",
        "DLY_Country",
        "DLY_PORT",
        "Booking_status",
        "Salesman_POR",
        "TEU_Booking",
    ]
    df1_cols = ["BKG_NO", "BKG_SHPR_CST_NO", "BKG_SHPR_CST_ENM"]

    df2 = pd.read_csv(
        ROOT / "2_2025.csv",
        encoding="utf-16",
        sep="\t",
        dtype=str,
        usecols=df2_cols,
        keep_default_na=False,
    )
    df2 = clean_columns(df2).drop_duplicates("BKG_NO", keep="first")

    for source, target in [
        ("POR_Country", "origin"),
        ("POR_PORT", "ori_port"),
        ("DLY_Country", "dest"),
        ("DLY_PORT", "dst_port"),
    ]:
        df2[target] = df2[source].map(norm_text)

    df2["team"] = [classify_team(o, d) for o, d in zip(df2["origin"], df2["dest"])]
    df2["basis_teu"] = to_number(df2["TEU_Booking"])
    df2["sales"] = df2["Salesman_POR"].map(lambda v: norm_text(v, "(미지정)"))

    df2 = df2.loc[
        (df2["team"] == TEAM)
        & df2["origin"].isin(ORIGINS)
        & df2["Booking_status"].astype(str).str.strip().eq("Normal")
        & df2["basis_teu"].gt(0)
    ].copy()

    df1 = pd.read_csv(
        ROOT / "1_2025.csv",
        encoding="utf-8-sig",
        dtype=str,
        usecols=df1_cols,
        keep_default_na=False,
    )
    df1 = clean_columns(df1).drop_duplicates("BKG_NO", keep="first")

    merged = df2.merge(df1, on="BKG_NO", how="left")
    merged["BKG_SHPR_CST_NO"] = merged["BKG_SHPR_CST_NO"].map(lambda v: norm_text(v, "(미확인)"))
    merged["BKG_SHPR_CST_ENM"] = merged["BKG_SHPR_CST_ENM"].map(lambda v: norm_text(v, "(미확인)"))

    keys = [
        "team",
        "origin",
        "ori_port",
        "dest",
        "dst_port",
        "BKG_SHPR_CST_NO",
        "BKG_SHPR_CST_ENM",
        "sales",
    ]
    return merged.groupby(keys, dropna=False)["basis_teu"].sum().reset_index()


def records_to_frame(records) -> pd.DataFrame:
    if not records:
        return pd.DataFrame()
    if isinstance(records, dict):
        columns = records.get("c") or records.get("columns")
        rows = records.get("r") or records.get("rows")
        dicts = records.get("d") or records.get("dicts") or {}
        if not isinstance(columns, list) or not isinstance(rows, list):
            return pd.DataFrame()
        expanded = []
        for row in rows:
            out = {}
            for idx, key in enumerate(columns):
                if idx >= len(row):
                    continue
                value = row[idx]
                if value is None:
                    continue
                dictionary = dicts.get(key)
                if isinstance(dictionary, list) and isinstance(value, int) and 0 <= value < len(dictionary):
                    value = dictionary[value]
                out[key] = value
            expanded.append(out)
        return pd.DataFrame(expanded)
    return pd.DataFrame(records)


def current_bsa_routes(data: dict) -> pd.DataFrame:
    bsa = records_to_frame(data.get("bsa", []))
    if bsa.empty:
        return pd.DataFrame()
    bsa = bsa.rename(columns={"POR_PORT": "ori_port", "DLY_PORT": "dst_port"})
    for col in ["team", "origin", "ori_port", "dest", "dst_port", "YYYYMM"]:
        if col not in bsa.columns:
            bsa[col] = ""
        bsa[col] = bsa[col].map(norm_text)
    bsa["teu_bsa"] = pd.to_numeric(bsa.get("teu_bsa", 0), errors="coerce").fillna(0.0)
    bsa = bsa.loc[
        (bsa["team"] == TEAM)
        & bsa["origin"].isin(ORIGINS)
        & bsa["YYYYMM"].isin(MONTHS)
        & bsa["teu_bsa"].gt(0)
    ].copy()
    keys = ["team", "origin", "ori_port", "dest", "dst_port", "YYYYMM"]
    return bsa.groupby(keys, dropna=False)["teu_bsa"].sum().reset_index()


def build_basis_lookup(basis: pd.DataFrame) -> dict[str, tuple[list[str], dict[tuple, pd.DataFrame]]]:
    levels = {
        "동일 POR port + DLY port": ["team", "origin", "ori_port", "dest", "dst_port"],
        "동일 도착 port": ["team", "origin", "dest", "dst_port"],
        "동일 도착국가": ["team", "origin", "dest"],
        "동일 선적국가": ["team", "origin"],
    }
    out: dict[str, tuple[list[str], dict[tuple, pd.DataFrame]]] = {}
    detail_cols = ["BKG_SHPR_CST_NO", "BKG_SHPR_CST_ENM", "sales"]
    for level, keys in levels.items():
        frame = basis.groupby(keys + detail_cols, dropna=False)["basis_teu"].sum().reset_index()
        lookup = {key if isinstance(key, tuple) else (key,): grp.reset_index(drop=True) for key, grp in frame.groupby(keys, dropna=False)}
        out[level] = (keys, lookup)
    return out


def allocate_bsa_to_sales(data: dict, basis: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    bsa_routes = current_bsa_routes(data)
    lookups = build_basis_lookup(basis)
    rows: list[dict] = []

    for _, route in bsa_routes.iterrows():
        selected_level = "배분근거 없음"
        selected_keys: list[str] = []
        selected_basis = pd.DataFrame()
        selected_total = 0.0

        for level, (keys, lookup) in lookups.items():
            key = tuple(route[k] for k in keys)
            group = lookup.get(key)
            total = float(group["basis_teu"].sum()) if group is not None else 0.0
            if total > 0:
                selected_level = level
                selected_keys = keys
                selected_basis = group
                selected_total = total
                break

        if selected_total <= 0:
            rows.append(
                {
                    "월": route["YYYYMM"],
                    "팀": route["team"],
                    "선적국가": route["origin"],
                    "선적포트": route["ori_port"],
                    "도착국가": route["dest"],
                    "도착포트": route["dst_port"],
                    "영업사원": "(배분근거 없음)",
                    "화주코드": "",
                    "화주명": "",
                    "2025_기준TEU": 0.0,
                    "2025_구간기준TEU": 0.0,
                    "배분비중": 0.0,
                    "구간월BSA": float(route["teu_bsa"]),
                    "배분BSA": float(route["teu_bsa"]),
                    "배분레벨": selected_level,
                    "배분키": "",
                }
            )
            continue

        key_desc = ", ".join(f"{k}={route[k]}" for k in selected_keys)
        for _, item in selected_basis.iterrows():
            share = float(item["basis_teu"]) / selected_total
            rows.append(
                {
                    "월": route["YYYYMM"],
                    "팀": route["team"],
                    "선적국가": route["origin"],
                    "선적포트": route["ori_port"],
                    "도착국가": route["dest"],
                    "도착포트": route["dst_port"],
                    "영업사원": item["sales"],
                    "화주코드": item["BKG_SHPR_CST_NO"],
                    "화주명": item["BKG_SHPR_CST_ENM"],
                    "2025_기준TEU": float(item["basis_teu"]),
                    "2025_구간기준TEU": selected_total,
                    "배분비중": share,
                    "구간월BSA": float(route["teu_bsa"]),
                    "배분BSA": float(route["teu_bsa"]) * share,
                    "배분레벨": selected_level,
                    "배분키": key_desc,
                }
            )

    detail = pd.DataFrame(rows)
    sales = (
        detail.groupby(["선적국가", "월", "영업사원"], dropna=False)["배분BSA"].sum().reset_index()
        if not detail.empty
        else pd.DataFrame(columns=["선적국가", "월", "영업사원", "배분BSA"])
    )
    return detail, sales


def build_current_sales_performance(data: dict) -> pd.DataFrame:
    shippers = records_to_frame(data.get("shipper", []))
    if shippers.empty:
        return pd.DataFrame()
    for col in ["team", "origin", "YYYYMM", "Salesman_POR", "BKG_SHPR_CST_NO"]:
        if col not in shippers.columns:
            shippers[col] = ""
        shippers[col] = shippers[col].map(norm_text)
    for col in ["w3_fst", "w3_norm_lst", "w3_route_hi_fst"]:
        if col not in shippers.columns:
            shippers[col] = 0.0
        shippers[col] = pd.to_numeric(shippers[col], errors="coerce").fillna(0.0)
    shippers["영업사원"] = shippers["Salesman_POR"].replace("", "(미지정)")
    shippers = shippers.loc[
        (shippers["team"] == TEAM)
        & shippers["origin"].isin(ORIGINS)
        & shippers["YYYYMM"].isin(MONTHS)
    ].copy()
    grouped = shippers.groupby(["origin", "YYYYMM", "영업사원"], dropna=False).agg(
        BKG=("w3_fst", "sum"),
        실선적=("w3_norm_lst", "sum"),
        구간별고수익=("w3_route_hi_fst", "sum"),
        화주수=("BKG_SHPR_CST_NO", lambda s: s[s.astype(str).str.strip().ne("")].nunique()),
    )
    return grouped.reset_index().rename(columns={"origin": "선적국가", "YYYYMM": "월"})


def build_sales_summary(perf: pd.DataFrame, allocated_sales: pd.DataFrame) -> pd.DataFrame:
    out = perf.merge(allocated_sales, on=["선적국가", "월", "영업사원"], how="outer")
    for col in ["BKG", "실선적", "구간별고수익", "화주수", "배분BSA"]:
        if col not in out.columns:
            out[col] = 0.0
        out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0.0)
    out["3주전_BKG_BSA"] = [pct_value(a, b) for a, b in zip(out["BKG"], out["배분BSA"])]
    out["고수익비중"] = [pct_value(a, b) for a, b in zip(out["구간별고수익"], out["BKG"])]
    out["실선적률"] = [pct_value(a, b) for a, b in zip(out["실선적"], out["BKG"])]
    out["총BKG"] = out.groupby(["선적국가", "영업사원"], dropna=False)["BKG"].transform("sum")
    out["총BSA"] = out.groupby(["선적국가", "영업사원"], dropna=False)["배분BSA"].transform("sum")
    return out.sort_values(["선적국가", "총BKG", "총BSA", "영업사원"], ascending=[True, False, False, True])


def build_destination_summary(data: dict) -> pd.DataFrame:
    monthly = records_to_frame(data.get("monthly", []))
    bsa = current_bsa_routes(data)
    if monthly.empty:
        return pd.DataFrame()
    for col in ["team", "origin", "dest", "YYYYMM"]:
        if col not in monthly.columns:
            monthly[col] = ""
        monthly[col] = monthly[col].map(norm_text)
    for col in ["w3_fst", "w3_norm_lst", "w3_route_hi_fst"]:
        if col not in monthly.columns:
            monthly[col] = 0.0
        monthly[col] = pd.to_numeric(monthly[col], errors="coerce").fillna(0.0)
    monthly = monthly.loc[
        (monthly["team"] == TEAM)
        & monthly["origin"].isin(ORIGINS)
        & monthly["YYYYMM"].isin(MONTHS)
    ].copy()
    perf = monthly.groupby(["origin", "dest", "YYYYMM"], dropna=False).agg(
        BKG=("w3_fst", "sum"),
        실선적=("w3_norm_lst", "sum"),
        구간별고수익=("w3_route_hi_fst", "sum"),
    ).reset_index()
    route_bsa = bsa.groupby(["origin", "dest", "YYYYMM"], dropna=False)["teu_bsa"].sum().reset_index()
    out = perf.merge(route_bsa, on=["origin", "dest", "YYYYMM"], how="outer")
    for col in ["BKG", "실선적", "구간별고수익", "teu_bsa"]:
        out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0.0)
    out["3주전_BKG_BSA"] = [pct_value(a, b) for a, b in zip(out["BKG"], out["teu_bsa"])]
    out["고수익비중"] = [pct_value(a, b) for a, b in zip(out["구간별고수익"], out["BKG"])]
    out["실선적률"] = [pct_value(a, b) for a, b in zip(out["실선적"], out["BKG"])]
    return out.rename(columns={"origin": "선적국가", "dest": "도착국가", "YYYYMM": "월"})


def aggregate_origin_rows(destination_summary: pd.DataFrame) -> pd.DataFrame:
    base = destination_summary.groupby(["선적국가", "월"], dropna=False).agg(
        BKG=("BKG", "sum"),
        실선적=("실선적", "sum"),
        구간별고수익=("구간별고수익", "sum"),
        BSA=("teu_bsa", "sum"),
    ).reset_index()
    base["3주전_BKG_BSA"] = [pct_value(a, b) for a, b in zip(base["BKG"], base["BSA"])]
    base["고수익비중"] = [pct_value(a, b) for a, b in zip(base["구간별고수익"], base["BKG"])]
    base["실선적률"] = [pct_value(a, b) for a, b in zip(base["실선적"], base["BKG"])]
    return base


def metric_values(frame: pd.DataFrame, label_col: str, label: str, origin: str, metric: str) -> list[float | None]:
    vals = []
    for month in MONTHS:
        match = frame.loc[(frame["선적국가"] == origin) & (frame[label_col] == label) & (frame["월"] == month)]
        vals.append(None if match.empty else match.iloc[0][metric])
    return vals


def origin_metric_values(frame: pd.DataFrame, origin: str, metric: str) -> list[float | None]:
    vals = []
    for month in MONTHS:
        match = frame.loc[(frame["선적국가"] == origin) & (frame["월"] == month)]
        vals.append(None if match.empty else match.iloc[0][metric])
    return vals


def write_metric_row(ws, row: int, col: int, label: str, values_by_metric: dict[str, list[float | None]], fills: dict[str, PatternFill]) -> None:
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.cell(row, col, label)
    ws.cell(row, col).fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(row, col).font = Font(bold=True)
    ws.cell(row, col).border = border
    ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")

    offset = 1
    for metric in ["3주전_BKG_BSA", "고수익비중", "실선적률"]:
        values = values_by_metric[metric]
        for idx, value in enumerate(values):
            cell = ws.cell(row, col + offset + idx, pct_text(value))
            cell.fill = fills[metric]
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        trend = ws.cell(row, col + offset + 3, trend_text(values))
        trend.fill = fills[metric]
        trend.font = Font(color="FF0000", bold=True)
        trend.border = border
        trend.alignment = Alignment(horizontal="center")
        offset += 4


def add_table_header(ws, row: int, col: int, origin: str, title: str = "선적지") -> None:
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fills = {
        "label": PatternFill("solid", fgColor="FFF2CC"),
        "3주전_BKG_BSA": PatternFill("solid", fgColor="FCE4D6"),
        "고수익비중": PatternFill("solid", fgColor="DDEBF7"),
        "실선적률": PatternFill("solid", fgColor="FCE4D6"),
    }
    ws.cell(row, col, title)
    ws.cell(row, col).fill = fills["label"]
    ws.cell(row, col).font = Font(bold=True)
    ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, col).border = border
    ws.merge_cells(start_row=row, start_column=col + 1, end_row=row, end_column=col + 4)
    ws.merge_cells(start_row=row, start_column=col + 5, end_row=row, end_column=col + 8)
    ws.merge_cells(start_row=row, start_column=col + 9, end_row=row, end_column=col + 12)
    headers = [
        ("3주전/BSA", "3주전_BKG_BSA", col + 1),
        ("고수익비중", "고수익비중", col + 5),
        ("실선적률", "실선적률", col + 9),
    ]
    for text, metric, start_col in headers:
        cell = ws.cell(row, start_col, text)
        cell.fill = fills[metric]
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        for c in range(start_col, start_col + 4):
            ws.cell(row, c).fill = fills[metric]
            ws.cell(row, c).border = border
    ws.cell(row + 1, col, origin)
    ws.cell(row + 1, col).fill = fills["label"]
    ws.cell(row + 1, col).font = Font(bold=True)
    ws.cell(row + 1, col).alignment = Alignment(horizontal="center")
    ws.cell(row + 1, col).border = border
    month_headers = [MONTH_LABELS[m] for m in MONTHS] + ["증감"]
    for block, metric in enumerate(["3주전_BKG_BSA", "고수익비중", "실선적률"]):
        start_col = col + 1 + block * 4
        for i, text in enumerate(month_headers):
            cell = ws.cell(row + 1, start_col + i, text)
            cell.fill = fills[metric]
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border


def write_dataframe_sheet(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for c, name in enumerate(df.columns, start=start_col):
        cell = ws.cell(start_row, c, name)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(r_idx, c_idx, value)
            cell.border = border
            if isinstance(value, float):
                cell.number_format = "#,##0.0000"
    ws.freeze_panes = ws.cell(start_row + 1, start_col).coordinate
    ws.auto_filter.ref = ws.dimensions
    for idx, name in enumerate(df.columns, start=start_col):
        width = min(max(12, len(str(name)) + 2), 42)
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_report(
    sales_summary: pd.DataFrame,
    destination_summary: pd.DataFrame,
    allocation_detail: pd.DataFrame,
    origin_summary: pd.DataFrame,
    validation: pd.DataFrame,
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "TH_VN_표"
    ws.sheet_view.showGridLines = False

    fills = {
        "3주전_BKG_BSA": PatternFill("solid", fgColor="FCE4D6"),
        "고수익비중": PatternFill("solid", fgColor="DDEBF7"),
        "실선적률": PatternFill("solid", fgColor="FCE4D6"),
    }

    start_cols = {"TH": 1, "VN": 16}
    for origin, start_col in start_cols.items():
        add_table_header(ws, 1, start_col, origin)
        row = 3
        values = {metric: origin_metric_values(origin_summary, origin, metric) for metric in fills}
        write_metric_row(ws, row, start_col, origin, values, fills)
        row += 2
        ws.cell(row, start_col, "* 도착지별").font = Font(bold=True)
        row += 1
        for dest in DEST_ROWS[origin]:
            values = {metric: metric_values(destination_summary, "도착국가", dest, origin, metric) for metric in fills}
            write_metric_row(ws, row, start_col, dest, values, fills)
            row += 1
        row += 2
        ws.cell(row, start_col, "* 영업사원별").font = Font(bold=True)
        row += 1
        origin_sales = (
            sales_summary.loc[sales_summary["선적국가"].eq(origin), ["영업사원", "총BKG", "총BSA"]]
            .drop_duplicates()
            .sort_values(["총BKG", "총BSA", "영업사원"], ascending=[False, False, True])
        )
        for sales in origin_sales["영업사원"]:
            values = {metric: metric_values(sales_summary, "영업사원", sales, origin, metric) for metric in fills}
            write_metric_row(ws, row, start_col, sales, values, fills)
            row += 1

    for c in range(1, 29):
        ws.column_dimensions[get_column_letter(c)].width = 8.5
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["P"].width = 16

    tidy = sales_summary.copy()
    tidy["3주전/BSA"] = tidy["3주전_BKG_BSA"].map(pct_text)
    tidy["고수익비중%"] = tidy["고수익비중"].map(pct_text)
    tidy["실선적률%"] = tidy["실선적률"].map(pct_text)
    tidy = tidy[
        [
            "선적국가",
            "월",
            "영업사원",
            "화주수",
            "BKG",
            "배분BSA",
            "3주전/BSA",
            "구간별고수익",
            "고수익비중%",
            "실선적",
            "실선적률%",
        ]
    ].sort_values(["선적국가", "영업사원", "월"])
    write_dataframe_sheet(wb.create_sheet("영업사원별_집계"), tidy)

    ws_basis = wb.create_sheet("BSA배분근거")
    notes = [
        ["배분 원칙", "2026년 3~5월 OBT 월별 BSA를 2025년 화주-구간 Normal TEU 실적 비중으로 배분했습니다."],
        ["정확 매칭", "team+선적국가+선적포트+도착국가+도착포트가 2025년에 있으면 동일 포트 구간 비중을 사용했습니다."],
        ["대체 매칭", "동일 포트 구간이 없으면 동일 도착포트, 동일 도착국가, 동일 선적국가 순서로 2025년 비중을 대체 적용했습니다."],
        ["집계", "화주별 배분BSA를 2025년 해당 화주 담당 영업사원 기준으로 합산했습니다."],
    ]
    for r, (k, v) in enumerate(notes, start=1):
        ws_basis.cell(r, 1, k).font = Font(bold=True)
        ws_basis.cell(r, 2, v)
    write_dataframe_sheet(ws_basis, allocation_detail, start_row=6)

    write_dataframe_sheet(wb.create_sheet("도착지별_검산"), destination_summary)
    write_dataframe_sheet(wb.create_sheet("검증"), validation)

    path = OUT_DIR / f"TH_VN_영업사원별_BSA배분_{DATASET_ID}.xlsx"
    wb.save(path)
    return path


def build_validation(bsa_routes: pd.DataFrame, allocation_detail: pd.DataFrame) -> pd.DataFrame:
    bsa_total = bsa_routes.groupby(["origin", "YYYYMM"], dropna=False)["teu_bsa"].sum().reset_index()
    alloc_total = (
        allocation_detail.groupby(["선적국가", "월"], dropna=False)["배분BSA"].sum().reset_index()
        if not allocation_detail.empty
        else pd.DataFrame(columns=["선적국가", "월", "배분BSA"])
    )
    out = bsa_total.merge(
        alloc_total,
        left_on=["origin", "YYYYMM"],
        right_on=["선적국가", "월"],
        how="outer",
    )
    out["선적국가"] = out["origin"].fillna(out["선적국가"])
    out["월"] = out["YYYYMM"].fillna(out["월"])
    out["원천BSA"] = pd.to_numeric(out["teu_bsa"], errors="coerce").fillna(0.0)
    out["배분BSA"] = pd.to_numeric(out["배분BSA"], errors="coerce").fillna(0.0)
    out["차이"] = out["원천BSA"] - out["배분BSA"]

    level = allocation_detail.groupby(["선적국가", "월", "배분레벨"], dropna=False)["배분BSA"].sum().reset_index()
    level_pivot = level.pivot_table(index=["선적국가", "월"], columns="배분레벨", values="배분BSA", aggfunc="sum", fill_value=0).reset_index()
    out = out[["선적국가", "월", "원천BSA", "배분BSA", "차이"]].merge(level_pivot, on=["선적국가", "월"], how="left")
    return out.sort_values(["선적국가", "월"])


def main() -> None:
    data = load_current_summary()
    basis = load_2025_basis()
    allocation_detail, allocated_sales = allocate_bsa_to_sales(data, basis)
    perf = build_current_sales_performance(data)
    sales_summary = build_sales_summary(perf, allocated_sales)
    destination_summary = build_destination_summary(data)
    origin_summary = aggregate_origin_rows(destination_summary)
    validation = build_validation(current_bsa_routes(data), allocation_detail)

    OUT_DIR.mkdir(exist_ok=True)
    allocation_csv = OUT_DIR / f"TH_VN_BSA배분근거_{DATASET_ID}.csv"
    summary_csv = OUT_DIR / f"TH_VN_영업사원별_집계_{DATASET_ID}.csv"
    allocation_detail.to_csv(allocation_csv, index=False, encoding="utf-8-sig")
    sales_summary.to_csv(summary_csv, index=False, encoding="utf-8-sig")
    xlsx = write_report(sales_summary, destination_summary, allocation_detail, origin_summary, validation)

    report = {
        "dataset_id": DATASET_ID,
        "months": MONTHS,
        "scope": {"team": TEAM, "origins": ORIGINS},
        "sources": {
            "current_summary": str(OUT_DIR / f"dashboard_summary_{DATASET_ID}.json"),
            "current_bsa": str(OUT_DIR / f"BSA_raw_monthly3W_{DATASET_ID}.csv"),
            "basis_2025_view1": str(ROOT / "1_2025.csv"),
            "basis_2025_view2": str(ROOT / "2_2025.csv"),
        },
        "allocation_rule": (
            "2026 monthly route BSA is allocated to shippers by 2025 Normal TEU basis. "
            "The match priority is exact POR port + DLY port route, same DLY port, "
            "same destination country, then same origin country. Shipper allocation is "
            "summed by the 2025 Salesman_POR."
        ),
        "outputs": {
            "xlsx": str(xlsx),
            "allocation_basis_csv": str(allocation_csv),
            "sales_summary_csv": str(summary_csv),
        },
        "validation": validation.to_dict("records"),
    }
    report_path = OUT_DIR / f"TH_VN_영업사원별_BSA배분_report_{DATASET_ID}.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"xlsx={xlsx}")
    print(f"allocation_basis_csv={allocation_csv}")
    print(f"sales_summary_csv={summary_csv}")
    print(f"report={report_path}")
    print(validation.to_string(index=False))


if __name__ == "__main__":
    main()
