from __future__ import annotations

import argparse
import importlib.util
import io
import json
import math
import re
import time
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import pandas as pd
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "output"

SALES_OWNER_SOURCE_ID = "1aGn2YyvKRx35mOsHLQAMaas6sa81LTOf4pNPugIUajg"
TARGET_SPREADSHEET_ID = "1YxZkwvoMaQXIEw07qUDZtCPDFZBf8GOZyr5knkxnLxo"
SALES_OWNER_INPUT_SHEET = "Sales_Owner_Input"

CURRENT_DATE_INT = 20260521
CURRENT_DATASET_ID = "20260521"
TEAM_FILTER = "OBT"
MISSING_SALES = "(\ubbf8\uc9c0\uc815)"
NO_BASIS_LABEL = "(no 2025 basis)"

MONTHS_2025 = {f"2025{m:02d}" for m in range(1, 13)}
Q1_2026 = {"202601", "202602", "202603"}
Q2_2026 = {"202604", "202605", "202606"}
Q2_PROGRESS_WEEKS = set(range(14, 20))
TARGET_MONTHS = MONTHS_2025 | Q1_2026 | Q2_2026
HIGH_TOKEN = "\uace0\uc218\uc775"

FISCAL_445 = {
    2025: (datetime(2024, 12, 29), [4, 4, 5, 4, 4, 5, 4, 4, 5, 4, 4, 6]),
    2026: (datetime(2026, 1, 4), [4, 4, 5, 4, 4, 5, 4, 4, 5, 4, 4, 5]),
    2027: (datetime(2027, 1, 3), [4, 4, 5, 4, 4, 5, 4, 4, 5, 4, 4, 5]),
}


# Target workbook tab -> source organization-chart sheet / column / row span.
SALES_OWNER_RANGES: dict[str, list[tuple[str, str, int, int]]] = {
    "CN_SHA": [("SHA", "H", 19, 24)],
    "CN_NKG": [("YZR", "F", 17, 19)],
    "CN_NBO": [("NBO", "G", 16, 18)],
    "CN_TAO": [("TAO", "G", 17, 20)],
    "CN_LYG": [("TAO", "J", 18, 18)],
    "CN_XGG": [("XGG", "G", 16, 18)],
    "CN_DLC": [("DLC", "F", 16, 18)],
    "CN_NNS": [("CAN", "G", 18, 20)],
    "CN_SHK_DCB": [("SZP", "G", 17, 20)],
    "CN_XMN": [("XMN", "F", 18, 20)],
    "HK": [("HK", "J", 22, 26)],
    "SG": [("SG", "G", 18, 21)],
    "TH": [("TH", "H", 21, 26)],
    "VN_SGN_CMP": [("VN", "E", 19, 22)],
    "VN_HPH": [("VN", "V", 19, 21)],
    "PH": [("PH", "H", 17, 20)],
    "IN": [("IN", "I", 22, 26), ("IN", "L", 21, 27), ("IN", "P", 22, 27), ("IN", "T", 21, 27)],
    "AE": [("AE", "F", 16, 20)],
    "JKT": [("ID", "F", 18, 24)],
    "SUB": [("ID", "Q", 20, 24)],
    "ID_out": [("ID", "G", 20, 24)],
    "PKG+PKW": [("MY", "N", 21, 23)],
    "PGU": [("MY", "R", 23, 23)],
    "PEN": [("MY", "W", 23, 23)],
}


MANUAL_ALIASES = {
    "JENNYDOU": "JENNY",
    "CRUISESONG": "CRUISE",
    "TERRYWANG": "TERRY",
    "COLINYANG": "COLIN",
    "BONOYUAN": "BONO",
    "CHLOEZHOU": "CHLOE",
    "SCOTTGAO": "SCOTT",
    "BILLGUAN": "BILL",
    "ROSSFU": "ROSS",
    "MARTINCHANG": "MARTIN1",
    "JASONWU": "JCWU",
    "JOELI": "CANID",
    "CHAMHM": "CHANHM",
    "PETERKONG": "KONG",
    "DARRENSIM": "DARREN",
    "ALEXLIN": "ALEX",
    "KELVINTAY": "KVIN",
    "ALVINANG": "ALVINA",
    "AKBAR": "AKBARA",
    "KEVIN": "KEVINYESAY",
    "RINATAN": "RINA",
    "EDWINKHU": "EDWIN",
    "JESSQAWOO": "JESSIQA",
    "TEHTL": "TLTEH",
    "WUXIAOCHEN": "ALEXWU",
    "CHICO": "CHICOYU",
    "HENRUHUANG": "HENRYHUANG",
    "WHZL": "LEOWANG",
    "MARTIN": "MARTIN1",
    "MICHAEL": "MICHAELWEI",
}

PREFERRED_ALIASES = {
    # ICC IDs in Global Network that are stale/short forms while salesman.csv
    # keeps the active customer owner under the expanded ID.
    "MANOJ": "ARYA",
    "PAPADA": "PAPHADA",
    "WINNIE": "WINNIESIA",
}


EXCLUDE_KEYS = {
    "SALES",
    "MARKETING",
    "SALESMARKETING",
    "SALESDEPARTMENT",
    "SALESDIVISION",
    "MKTSLS",
    "SALESCS",
    "SALESC",
    "CS",
    "OBSALES",
    "IBSALES",
    "BSALES",
    "TOTAL",
    "NUMBER",
    "CONFIDENTIAL",
    "OUTPORT",
}


RAW_KEYS = (
    "lst_2025",
    "w3_2025",
    "bsa_2025",
    "w3_norm_lst_2025",
    "hi_w3_2025",
    "w3_q1",
    "bsa_q1",
    "w3_norm_lst_q1",
    "hi_w3_q1",
    "w3_q2_progress",
    "bsa_q2_progress",
    "w3_norm_lst_q2_progress",
    "hi_w3_q2_progress",
)


@dataclass(frozen=True)
class OwnerEntry:
    target_tab: str
    source_sheet: str
    source_cell: str
    input_name: str
    sales_key: str
    resolved_sales: str
    match_status: str
    source_type: str
    customer_count: int


def load_module(name: str, path: Path) -> Any:
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def target_builder() -> Any:
    return load_module("target_builder", ROOT / "scripts" / "create_per_origin_target_workbook.py")


def clean_text(value: Any, fallback: str = "") -> str:
    if value is None:
        return fallback
    try:
        if pd.isna(value):
            return fallback
    except TypeError:
        pass
    text = str(value).strip()
    return fallback if text.lower() in {"", "nan", "none", "nat"} else text


def norm_key(value: object) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def to_number(series: pd.Series) -> pd.Series:
    return pd.to_numeric(
        series.fillna("").astype(str).str.replace(",", "", regex=False).str.strip(),
        errors="coerce",
    ).fillna(0.0)


def safe_ratio(num: Any, den: Any) -> float | None:
    try:
        n = float(num)
        d = float(den)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(n) or not math.isfinite(d) or d == 0:
        return None
    return n / d


def fiscal_week_lookup() -> dict[str, int]:
    out: dict[str, int] = {}
    for year, (first_sun, pattern) in FISCAL_445.items():
        week_no = 1
        for weeks_in_month in pattern:
            for _ in range(weeks_in_month):
                out[(first_sun + timedelta(weeks=week_no - 1)).strftime("%Y-%m-%d")] = week_no
                week_no += 1
    return out


FISCAL_WEEK_BY_START = fiscal_week_lookup()


def week_start_key(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    if isinstance(value, (datetime, pd.Timestamp)):
        return value.strftime("%Y-%m-%d")
    match = re.search(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", text)
    if not match:
        return ""
    year, month, day = (int(part) for part in match.groups())
    return f"{year:04d}-{month:02d}-{day:02d}"


def fiscal_week(value: Any) -> int | None:
    key = week_start_key(value)
    if not key:
        return None
    return FISCAL_WEEK_BY_START.get(key)


def colnum(letter: str) -> int:
    out = 0
    for ch in letter:
        out = out * 26 + ord(ch.upper()) - 64
    return out


def q(title: str) -> str:
    return "'" + title.replace("'", "''") + "'"


def get_creds() -> Credentials:
    creds_dir = ROOT.parent / ".gdrive-mcp"
    credentials_path = creds_dir / "credentials.json"
    token_path = creds_dir / "token.json"
    installed = json.loads(credentials_path.read_text(encoding="utf-8-sig"))["installed"]
    token = json.loads(token_path.read_text(encoding="utf-8-sig"))
    creds = Credentials(
        token=token.get("access_token"),
        refresh_token=token["refresh_token"],
        token_uri=installed["token_uri"],
        client_id=installed["client_id"],
        client_secret=installed["client_secret"],
    )
    creds.refresh(Request())
    return creds


def with_backoff(call: Any) -> Any:
    for attempt in range(7):
        try:
            return call.execute()
        except Exception as exc:  # noqa: BLE001
            if "429" not in str(exc) or attempt == 6:
                raise
            time.sleep((2**attempt) * 5)
    raise RuntimeError("unreachable")


def export_sheet(drive: Any, file_id: str, filename: str) -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = OUT_DIR / filename
    request = drive.files().export_media(
        fileId=file_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    with io.FileIO(out, "wb") as fh:
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
    return out


def split_names(value: object) -> list[str]:
    if value is None:
        return []
    text = str(value).replace("\xa0", " ").replace("*", "").strip()
    text = re.sub(
        r"\([^)]*(?:DRIV|DRIVER|IT|HR|IB\.|OB\.|additional post|Messenger|Maid)[^)]*\)",
        "",
        text,
        flags=re.I,
    )
    parts = re.split(r"\s*/\s*|\s*,\s*|\n+|\s{2,}", text)
    names: list[str] = []
    for part in parts:
        part = re.sub(r"\([^)]*\)", "", part).strip()
        part = re.sub(r"\s+", " ", part)
        key = norm_key(part)
        if not key or key.isdigit() or key in EXCLUDE_KEYS or len(key) <= 1:
            continue
        names.append(part)
    return names


def workbook_values(path: Path, *, read_only: bool = True) -> dict[str, list[list[Any]]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=read_only)
    return {
        ws.title: [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
        for ws in wb.worksheets
    }


def read_existing_target_input(target_path: Path, mod: Any) -> dict[str, tuple[float, float, float, str]]:
    wb = openpyxl.load_workbook(target_path, data_only=True, read_only=True)
    if mod.INPUT_SHEET not in wb.sheetnames:
        return {}
    ws = wb[mod.INPUT_SHEET]
    existing: dict[str, tuple[float, float, float, str]] = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        tab = clean_text(row[0] if row else "")
        if not tab:
            continue
        booking = mod.clean_number(row[1]) or 0.0
        lifting = mod.clean_number(row[3]) or 0.0
        hp = mod.clean_number(row[5]) or 0.0
        memo = "" if len(row) < 8 or row[7] is None else str(row[7])
        existing[tab] = (booking, lifting, hp, memo)
    return existing


def active_salesman_mapping() -> tuple[pd.DataFrame, pd.Series, dict[str, int]]:
    path = ROOT / "saleman.csv"
    if not path.exists():
        path = ROOT / "salesman.csv"
    if not path.exists():
        raise FileNotFoundError("Missing saleman.csv or salesman.csv")
    sales = pd.read_csv(path, dtype=str, encoding="utf-8-sig", low_memory=False)
    for col in ["COUNTRY", "PORT", "SALESMAN_NO", "CUSTOMER_NO", "SALES_START_DATE", "SALES_END_DATE"]:
        sales[col] = sales[col].fillna("").astype(str).str.strip()
    start = pd.to_numeric(sales["SALES_START_DATE"].str.replace(".0", "", regex=False), errors="coerce")
    end = pd.to_numeric(sales["SALES_END_DATE"].str.replace(".0", "", regex=False), errors="coerce")
    active = sales.loc[(start <= CURRENT_DATE_INT) & (end >= CURRENT_DATE_INT)].copy()
    active["CUSTOMER_NO_KEY"] = active["CUSTOMER_NO"].str.upper()
    active["SALESMAN_NO"] = active["SALESMAN_NO"].map(clean_text)
    active = active.loc[active["CUSTOMER_NO_KEY"].ne("")]
    active = active.drop_duplicates("CUSTOMER_NO_KEY", keep="first")
    customer_to_sales = active.set_index("CUSTOMER_NO_KEY")["SALESMAN_NO"]
    counts = active.groupby("SALESMAN_NO", dropna=False)["CUSTOMER_NO_KEY"].nunique().to_dict()
    return active, customer_to_sales, {clean_text(k, MISSING_SALES): int(v) for k, v in counts.items()}


def person_key(value: object) -> str:
    text = str(value or "")
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"\b(MR|MS|MRS|MISS|MDM|CAPT)\.?\b", " ", text, flags=re.I)
    return norm_key(text)


def id_key(value: object) -> str:
    text = clean_text(value)
    if "@" in text:
        text = text.split("@", 1)[0]
    if text in {"-", "N/A", "NA"}:
        return ""
    return norm_key(text)


def resolve_sales_from_keys(keys: list[str], available_by_key: dict[str, str]) -> tuple[str, str]:
    keys = [key for key in dict.fromkeys(keys) if key and key not in {"NA", "N", "A"}]
    for key in keys:
        alias = PREFERRED_ALIASES.get(key)
        if alias and norm_key(alias) in available_by_key:
            return available_by_key[norm_key(alias)], f"preferred_alias:{key}->{alias}"
    for key in keys:
        if key in available_by_key:
            return available_by_key[key], f"exact:{key}"
    for key in keys:
        alias = MANUAL_ALIASES.get(key)
        if alias and norm_key(alias) in available_by_key:
            return available_by_key[norm_key(alias)], f"alias:{key}->{alias}"

    candidates: list[str] = []
    for key in keys:
        for avail_key, name in available_by_key.items():
            if len(avail_key) >= 5 and len(key) >= 5 and (
                key.startswith(avail_key)
                or avail_key.startswith(key)
                or key.endswith(avail_key)
                or avail_key.endswith(key)
            ):
                candidates.append(name)
    candidates = sorted(set(candidates))
    if len(candidates) == 1:
        return candidates[0], f"prefix:{','.join(keys)}"
    fallback = keys[0] if keys else ""
    return fallback, "unmatched"


CONTACT_POINT_SHEETS = [
    "JP",
    "S.CN",
    "N.CN",
    "TH ",
    "TW",
    "HK",
    "PH",
    "VN",
    "SG",
    "MY",
    "BD",
    "IN",
    "LK",
    "AE+ME",
    "PK",
    "KH",
    "TZ,KE",
    "ID",
    "US",
    "MX",
    "RU",
]


SHEET_DEFAULT_TARGET = {
    "TH ": "TH",
    "TW": "TW",
    "HK": "HK",
    "PH": "PH",
    "VN": "VN_SGN_CMP",
    "SG": "SG",
    "MY": "PKG+PKW",
    "BD": "BD",
    "IN": "IN",
    "LK": "LK",
    "AE+ME": "AE",
    "PK": "PK",
    "KH": "KH",
    "TZ,KE": "TZ",
    "US": "US",
    "MX": "MX",
    "RU": "RU",
}


PORT_TARGET_MAP = {
    "SHA": "CN_SHA",
    "NKG": "CN_NKG",
    "YZR": "CN_NKG",
    "XGG": "CN_XGG",
    "DLC": "CN_DLC",
    "TAO": "CN_TAO",
    "LYG": "CN_LYG",
    "NBO": "CN_NBO",
    "SZP": "CN_SHK_DCB",
    "SHK": "CN_SHK_DCB",
    "DCB": "CN_SHK_DCB",
    "CAN": "CN_NNS",
    "NNS": "CN_NNS",
    "XMN": "CN_XMN",
    "KEL": "TW",
    "TPE": "TW",
    "TW": "TW",
    "HKG": "HK",
    "HK": "HK",
    "SIN": "SG",
    "SG": "SG",
    "BKK": "TH",
    "TH": "TH",
    "LCB": "TH",
    "SGN": "VN_SGN_CMP",
    "CMP": "VN_SGN_CMP",
    "HPH": "VN_HPH",
    "JKT": "JKT",
    "SUB": "SUB",
    "SRG": "ID_out",
    "BDG": "ID_out",
    "BLW": "ID_out",
    "BKS": "ID_out",
    "PJG": "ID_out",
    "BTM": "ID_out",
    "PKG": "PKG+PKW",
    "PKW": "PKG+PKW",
    "PEN": "PEN",
    "PGU": "PGU",
    "IN": "IN",
    "NSA": "IN",
    "DELHI": "IN",
    "DELHIINLANDOFFICE": "IN",
    "MAA": "IN",
    "CCU": "IN",
    "BLR": "IN",
    "COK": "IN",
    "TUT": "IN",
    "CMB": "LK",
    "KHI": "PK",
    "PK": "PK",
    "UAE": "AE",
    "AE": "AE",
    "DXB": "AE",
    "JEA": "AE",
    "SAUDI": "SA",
    "SA": "SA",
    "JED": "SA",
    "RUH": "SA",
    "KSP": "BH",
    "BAH": "BH",
    "MBA": "KE",
    "KE": "KE",
    "DAR": "TZ",
    "TZ": "TZ",
    "ALY": "EG",
    "EG": "EG",
    "ZLO": "MX",
    "LZO": "MX",
    "MX": "MX",
}


def contact_sheet_rows(ws: Any, max_rows: int = 1100, max_cols: int = 40) -> list[list[str]]:
    return [
        [clean_text(cell) for cell in row]
        for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True)
    ]


def contact_header(row: list[str]) -> dict[str, int | None] | None:
    keys = [norm_key(value) for value in row]
    name_idx = next((i for i, key in enumerate(keys) if key in {"NAME", "FULLNAME"} or key.endswith("NAME")), None)
    if name_idx is not None and any(token in keys[name_idx] for token in {"EMAIL", "GROUP"}):
        name_idx = None
    if name_idx is None and "PORT" in keys and "ORGANIZATION" in keys and ("ICCID" in keys or "ICC" in keys):
        # Some Global Network blocks label the person-name column as "Port";
        # the actual location is held in the adjacent Organization column.
        port_candidate = keys.index("PORT")
        org_candidate = keys.index("ORGANIZATION")
        if port_candidate < org_candidate:
            name_idx = port_candidate
    if name_idx is None:
        return None
    has_context = any(
        key in keys or any(item.startswith(key) for item in keys)
        for key in {"AREA", "DEPARTMENT", "DEPT", "JOB", "ORGANIZATION", "PORT", "POSITION", "ROLE"}
    )
    if not has_context:
        return None

    def idx(*names: str) -> int | None:
        for name in names:
            if name in keys:
                return keys.index(name)
            for i, key in enumerate(keys):
                if key.startswith(name):
                    return i
        return None

    icc_idx = idx("ICCID", "ICC")
    next_header = norm_key(row[name_idx + 1]) if name_idx + 1 < len(row) else ""
    name_extra_idx = name_idx + 1 if icc_idx is not None and name_idx + 1 < icc_idx and not next_header else None
    port_idx = idx("AREA", "PORT", "ORGANIZATION")
    if name_idx is not None and port_idx == name_idx:
        alternate_port = idx("ORGANIZATION", "AREA")
        port_idx = alternate_port if alternate_port != name_idx else None
    return {
        "port": port_idx,
        "dept": idx("DEPARTMENT", "DEPT"),
        "name": name_idx,
        "name_extra": name_extra_idx,
        "icc": icc_idx,
        "job": idx("JOB", "JOBDESCRIPTION", "ROLE", "POSITION"),
    }


def contact_saleslike(dept: str, job: str) -> bool:
    dept_u = dept.upper().replace("\n", " ")
    job_u = job.upper().replace("\n", " ")
    if any(token in dept_u for token in ["SALES", "MARKET", "MARKETING", "MKTG"]):
        return True
    if "BUSINESS TEAM" in dept_u:
        return True
    if dept_u.strip() == "BUSINESS" and any(token in job_u for token in ["SALES", "MARKET", "MARKETING"]):
        return True
    if any(token in dept_u for token in ["MANAGEMENT", "GENERAL MANAGER", "DIRECTOR", "OPERATION", "DOCUMENT", "ACCOUNT", "FINANCE", "HR", "ADMIN"]):
        return False
    if re.search(r"\bSLS\b", job_u):
        return True
    return any(token in job_u for token in ["SALES", "MARKET", "MARKETING"])


def contact_port_key(value: str) -> str:
    text = value.replace("\n", " ").strip()
    if not text:
        return ""
    first = text.split()[0]
    return norm_key(first)


def contact_name(row: list[str], header: dict[str, int | None]) -> str:
    name_idx = header["name"]
    if name_idx is None:
        return ""
    parts = [clean_text(row[name_idx]) if name_idx < len(row) else ""]
    extra_idx = header.get("name_extra")
    if extra_idx is not None and extra_idx < len(row):
        parts.append(clean_text(row[extra_idx]))
    return " ".join(part for part in parts if part).strip()


def looks_like_contact_name(value: str) -> bool:
    text = clean_text(value)
    if not text:
        return False
    upper = text.upper()
    if re.search(r"\b(ADDRESS|TEL|FAX|HTTP|WWW|EMAIL|E-MAIL|UPDATED)\b", upper):
        return False
    if ":" in text or re.search(r"\d", text):
        return False
    return True


def load_owner_entries(
    owner_path: Path,
    active_sales: pd.DataFrame,
    customer_counts: dict[str, int],
) -> dict[str, list[OwnerEntry]]:
    wb = openpyxl.load_workbook(owner_path, data_only=True, read_only=True)
    available = sorted({clean_text(x) for x in active_sales["SALESMAN_NO"].dropna() if clean_text(x)})
    available_by_key = {norm_key(name): name for name in available if norm_key(name)}
    parsed: dict[str, list[OwnerEntry]] = {}
    seen: set[tuple[str, str]] = set()
    for sheet in CONTACT_POINT_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        header: dict[str, int | None] | None = None
        current_port = ""
        current_dept = ""
        current_target = SHEET_DEFAULT_TARGET.get(sheet)
        for row_num, row in enumerate(contact_sheet_rows(wb[sheet]), start=1):
            found = contact_header(row)
            if found:
                header = found
                current_port = ""
                current_dept = ""
                current_target = SHEET_DEFAULT_TARGET.get(sheet)
                continue
            if header is None:
                continue

            port_idx = header.get("port")
            dept_idx = header.get("dept")
            job_idx = header.get("job")
            icc_idx = header.get("icc")
            port_text = clean_text(row[port_idx]) if port_idx is not None and port_idx < len(row) else ""
            dept = clean_text(row[dept_idx]) if dept_idx is not None and dept_idx < len(row) else ""
            job = clean_text(row[job_idx]) if job_idx is not None and job_idx < len(row) else ""
            icc = clean_text(row[icc_idx]) if icc_idx is not None and icc_idx < len(row) else ""
            name = contact_name(row, header)

            port_key = contact_port_key(port_text)
            if port_key in PORT_TARGET_MAP:
                current_port = port_key
                current_target = PORT_TARGET_MAP[port_key]
            if dept:
                current_dept = dept
            if not name or norm_key(name) == "NAME" or len(norm_key(name)) < 2 or not looks_like_contact_name(name):
                continue

            target_tab = PORT_TARGET_MAP.get(current_port, current_target)
            if not target_tab or not contact_saleslike(current_dept, job):
                continue

            keys = [id_key(icc), person_key(name), norm_key(name)]
            resolved, status = resolve_sales_from_keys(keys, available_by_key)
            dedupe_key = (target_tab, resolved or name)
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            entry = OwnerEntry(
                target_tab=target_tab,
                source_sheet=sheet,
                source_cell=f"R{row_num}",
                input_name=name,
                sales_key="/".join(key for key in keys if key),
                resolved_sales=resolved,
                match_status=status,
                source_type="contact_point",
                customer_count=customer_counts.get(resolved, 0),
            )
            parsed.setdefault(target_tab, []).append(entry)
    return parsed


def tab_key(origin: object, ori_port: object) -> str:
    origin_text = clean_text(origin, "UNKNOWN")
    port = clean_text(ori_port, "UNKNOWN")
    if origin_text == "CN":
        return "CN_SHK_DCB" if port in {"SHK", "DCB"} else f"CN_{port}"
    if origin_text == "VN":
        if port in {"SGN", "CMP"}:
            return "VN_SGN_CMP"
        if port == "HPH":
            return "VN_HPH"
        return f"VN_{port}"
    if origin_text == "ID":
        if port == "JKT":
            return "JKT"
        if port == "SUB":
            return "SUB"
        return "ID_out"
    if origin_text == "MY":
        if port in {"PKG", "PKW"}:
            return "PKG+PKW"
        if port == "PEN":
            return "PEN"
        if port == "PGU":
            return "PGU"
        return "MY_out"
    return origin_text


def classify_team(origin: object, dest: object) -> str:
    o = clean_text(origin)
    d = clean_text(dest)
    if o not in ("KR", "JP") and d != "KR":
        return "OBT"
    if o == "KR" and d != "JP":
        return "EST"
    if o != "JP" and d == "KR":
        return "IST"
    return "JBT"


def ensure_2025_cache() -> Path:
    out = OUT_DIR / "_cache_2025.parquet"
    if out.exists():
        return out
    mod = load_module("build_2025_bsa_shipper_sheet", ROOT / "scripts" / "build_2025_bsa_shipper_sheet.py")
    df = mod.add_dashboard_fields(mod.build_booking_frame())
    if "Lead_time (BKG_Sche)" not in df.columns and "Lead_time_BKG_Sche" in df.columns:
        df["Lead_time (BKG_Sche)"] = df["Lead_time_BKG_Sche"]
    df.to_parquet(out, index=False)
    return out


def first_existing(columns: list[str], candidates: list[str]) -> str | None:
    for candidate in candidates:
        if candidate in columns:
            return candidate
    return None


def high_flag_column(columns: list[str]) -> str | None:
    for col in columns:
        esc = col.encode("unicode_escape").decode()
        if "\\uace0/\\uc800" == esc:
            return col
    tagged = []
    for col in columns:
        esc = col.encode("unicode_escape").decode()
        if "\\ud0dc\\uadf8" in esc or "\\uace0\\uc218\\uc775" in esc:
            tagged.append(col)
    return tagged[0] if tagged else None


def normalize_booking_frame(path: Path, customer_to_sales: pd.Series) -> pd.DataFrame:
    frame = pd.read_parquet(path)
    cols = list(frame.columns)
    rename_map = {
        "BKG_NO": "bkg_no",
        "BKG_SHPR_CST_NO": "shipper_code",
        "BKG_SHPR_CST_ENM": "shipper_name",
        "POR_CTR_CD": "origin",
        "POR_PLC_CD": "ori_port",
        "DLY_CTR_CD": "dest",
        "DLY_PLC_CD": "dst_port",
        "FST_TEU": "fst",
        "LST_Status": "status",
        "LST_TEU": "lst",
        "YYYYMM": "yyyymm",
        "Salesman_POR": "legacy_sales",
    }
    lead_col = first_existing(cols, ["Lead_time (BKG_Sche)", "Lead_time_BKG_Sche"])
    week_col = first_existing(cols, ["week_start_date", "week_start_date_key"])
    high_col = high_flag_column(cols)
    usecols = [c for c in rename_map if c in frame.columns]
    if lead_col:
        usecols.append(lead_col)
    if week_col:
        usecols.append(week_col)
    if high_col:
        usecols.append(high_col)
    out = frame.loc[:, usecols].copy()
    out = out.rename(columns={k: v for k, v in rename_map.items() if k in out.columns})
    if lead_col:
        out = out.rename(columns={lead_col: "lead_time"})
    else:
        out["lead_time"] = ""
    if week_col:
        out = out.rename(columns={week_col: "week_start"})
    else:
        out["week_start"] = ""
    if high_col:
        out = out.rename(columns={high_col: "high_source"})
    else:
        out["high_source"] = ""
    for col in ["bkg_no", "shipper_code", "shipper_name", "origin", "ori_port", "dest", "dst_port", "status", "yyyymm", "lead_time", "week_start", "legacy_sales", "high_source"]:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].map(clean_text)
    out["fst"] = to_number(out.get("fst", pd.Series(dtype=str)))
    out["lst"] = to_number(out.get("lst", pd.Series(dtype=str)))
    out["shipper_code_key"] = out["shipper_code"].str.upper()
    mapped = out["shipper_code_key"].map(customer_to_sales)
    out["sales"] = mapped.map(clean_text).replace("", np.nan).fillna(MISSING_SALES)
    out["mapped_current_owner"] = mapped.notna() & mapped.map(clean_text).ne("")
    out["team"] = [classify_team(o, d) for o, d in zip(out["origin"], out["dest"])]
    out["tab"] = [tab_key(o, p) for o, p in zip(out["origin"], out["ori_port"])]
    out["is_high"] = out["high_source"].str.contains(HIGH_TOKEN, regex=False, na=False)
    out["fiscal_week"] = out["week_start"].map(fiscal_week)
    out = out.loc[out["yyyymm"].isin(TARGET_MONTHS) & out["team"].eq(TEAM_FILTER)].copy()
    keep = [
        "bkg_no",
        "shipper_code",
        "shipper_code_key",
        "shipper_name",
        "origin",
        "ori_port",
        "dest",
        "dst_port",
        "fst",
        "status",
        "lst",
        "yyyymm",
        "lead_time",
        "fiscal_week",
        "legacy_sales",
        "sales",
        "mapped_current_owner",
        "team",
        "tab",
        "is_high",
    ]
    return out[keep]


def load_booking(customer_to_sales: pd.Series) -> pd.DataFrame:
    paths = [ensure_2025_cache(), OUT_DIR / f"_cache_{CURRENT_DATASET_ID}.parquet"]
    missing = [str(path) for path in paths if not path.exists()]
    if missing:
        raise FileNotFoundError(", ".join(missing))
    frames = [normalize_booking_frame(path, customer_to_sales) for path in paths]
    return pd.concat(frames, ignore_index=True)


def load_bsa(week_filter: set[int] | None = None) -> pd.DataFrame:
    path = OUT_DIR / f"BSA_raw_monthly3W_{CURRENT_DATASET_ID}.csv"
    bsa = pd.read_csv(path, dtype=str, encoding="utf-8-sig", low_memory=False)
    bsa = bsa.rename(
        columns={
            "POR_Country": "origin",
            "POR_PORT": "ori_port",
            "DLY_Country": "dest",
            "DLY_PORT": "dst_port",
            "YYYYMM": "yyyymm",
            "TEU_BSA (Actual)": "route_bsa",
        }
    )
    for col in ["origin", "ori_port", "dest", "dst_port", "yyyymm"]:
        bsa[col] = bsa[col].map(clean_text)
    if week_filter is not None:
        bsa["ww"] = pd.to_numeric(bsa.get("WW", ""), errors="coerce")
        bsa = bsa.loc[bsa["ww"].isin(week_filter)].copy()
    bsa["team"] = bsa.get("team", bsa.get("Sales Team", "")).map(clean_text).str.upper()
    bsa["route_bsa"] = to_number(bsa["route_bsa"])
    bsa["tab"] = [tab_key(o, p) for o, p in zip(bsa["origin"], bsa["ori_port"])]
    bsa = bsa.loc[bsa["yyyymm"].isin(TARGET_MONTHS) & bsa["team"].eq(TEAM_FILTER) & bsa["route_bsa"].gt(0)].copy()
    keys = ["yyyymm", "tab", "team", "origin", "ori_port", "dest", "dst_port"]
    return bsa.groupby(keys, dropna=False)["route_bsa"].sum().reset_index()


def target_origins_from_workbook(path: Path, mod: Any) -> list[str]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    support = {mod.README_SHEET, mod.INPUT_SHEET, mod.SUMMARY_SHEET, SALES_OWNER_INPUT_SHEET}
    return [name for name in wb.sheetnames if name not in support]


def candidate_sales_by_origin(booking: pd.DataFrame) -> dict[str, list[str]]:
    metric_rows = booking.loc[
        booking["mapped_current_owner"]
        & booking["sales"].ne(MISSING_SALES)
        & (
            (booking["yyyymm"].isin(MONTHS_2025) & booking["lst"].gt(0))
            | (booking["yyyymm"].isin(Q1_2026) & booking["fst"].gt(0))
        )
    ]
    grouped = (
        metric_rows.groupby(["tab", "sales"], dropna=False)
        .agg(lst_2025=("lst", lambda s: float(s[metric_rows.loc[s.index, "yyyymm"].isin(MONTHS_2025)].sum())))
        .reset_index()
    )
    out: dict[str, list[str]] = {}
    if grouped.empty:
        return out
    sort_basis = (
        booking.loc[booking["mapped_current_owner"] & booking["yyyymm"].isin(MONTHS_2025)]
        .groupby(["tab", "sales"], dropna=False)["lst"]
        .sum()
        .reset_index(name="basis")
    )
    grouped = grouped.merge(sort_basis, on=["tab", "sales"], how="left")
    grouped["basis"] = grouped["basis"].fillna(0.0)
    for tab, group in grouped.sort_values(["tab", "basis", "sales"], ascending=[True, False, True]).groupby("tab", sort=False):
        out[str(tab)] = [str(s) for s in group["sales"].tolist() if str(s)]
    return out


def complete_owner_entries(
    origins: list[str],
    explicit: dict[str, list[OwnerEntry]],
    fallback_sales: dict[str, list[str]],
    customer_counts: dict[str, int],
) -> dict[str, list[OwnerEntry]]:
    out: dict[str, list[OwnerEntry]] = {}
    for origin in origins:
        entries = explicit.get(origin, [])
        if entries:
            out[origin] = entries
            continue
        out[origin] = [
            OwnerEntry(
                target_tab=origin,
                source_sheet="saleman.csv",
                source_cell="CUSTOMER_NO",
                input_name=sales,
                sales_key=norm_key(sales),
                resolved_sales=sales,
                match_status="active_customer_mapping",
                source_type="fallback_active_mapping",
                customer_count=customer_counts.get(sales, 0),
            )
            for sales in fallback_sales.get(origin, [])
        ]
    return out


def allowed_sales_by_origin(owner_entries: dict[str, list[OwnerEntry]]) -> dict[str, set[str]]:
    return {
        origin: {entry.resolved_sales for entry in entries if entry.resolved_sales and entry.match_status != "unmatched"}
        for origin, entries in owner_entries.items()
    }


def metric_sum(booking: pd.DataFrame, mask: pd.Series, value_col: str) -> pd.DataFrame:
    keys = ["tab", "sales", "yyyymm"]
    return (
        booking.loc[mask, keys + [value_col]]
        .groupby(keys, dropna=False)[value_col]
        .sum()
        .reset_index()
        .rename(columns={value_col: "value"})
    )


def scoped_booking(booking: pd.DataFrame, allowed: dict[str, set[str]]) -> pd.DataFrame:
    if not allowed:
        return booking.iloc[0:0].copy()
    mask = pd.Series(False, index=booking.index)
    for origin, sales in allowed.items():
        if not sales:
            continue
        mask |= booking["tab"].eq(origin) & booking["sales"].isin(sales)
    return booking.loc[mask].copy()


def allocate_bsa_scoped(bsa: pd.DataFrame, booking: pd.DataFrame, allowed: dict[str, set[str]]) -> pd.DataFrame:
    basis_rows = scoped_booking(booking, allowed)
    basis_rows = basis_rows.loc[
        basis_rows["yyyymm"].isin(MONTHS_2025)
        & basis_rows["status"].eq("Normal")
        & basis_rows["lst"].gt(0)
    ].copy()
    route_keys = ["tab", "team", "origin", "ori_port", "dest", "dst_port"]
    grouped = (
        basis_rows.groupby(route_keys + ["sales"], dropna=False)["lst"]
        .sum()
        .reset_index()
        .rename(columns={"lst": "basis_lST_TEU"})
    )
    lookups: dict[tuple[Any, ...], pd.DataFrame] = {}
    for key, group in grouped.groupby(route_keys, dropna=False):
        key_tuple = key if isinstance(key, tuple) else (key,)
        lookups[key_tuple] = group[["sales", "basis_lST_TEU"]].reset_index(drop=True)

    rows: list[dict[str, Any]] = []
    for route in bsa.itertuples(index=False):
        route_dict = route._asdict()
        origin = route_dict["tab"]
        sales_scope = allowed.get(origin, set())
        if not sales_scope:
            continue
        key = tuple(route_dict[k] for k in route_keys)
        group = lookups.get(key)
        if group is None or group.empty:
            continue
        group = group.loc[group["sales"].isin(sales_scope)]
        total = float(group["basis_lST_TEU"].sum())
        if total <= 0:
            continue
        route_bsa = float(route_dict["route_bsa"])
        for item in group.itertuples(index=False):
            basis = float(item.basis_lST_TEU)
            rows.append(
                {
                    "tab": origin,
                    "sales": item.sales,
                    "yyyymm": route_dict["yyyymm"],
                    "value": route_bsa * basis / total,
                }
            )
    return pd.DataFrame(rows, columns=["tab", "sales", "yyyymm", "value"])


def build_metric_frames(
    booking: pd.DataFrame,
    bsa: pd.DataFrame,
    q2_progress_bsa: pd.DataFrame,
    allowed: dict[str, set[str]],
) -> dict[str, pd.DataFrame]:
    scoped = scoped_booking(booking, allowed)
    lst = metric_sum(scoped, scoped["status"].eq("Normal") & scoped["lst"].gt(0), "lst")
    w3_mask = scoped["lead_time"].eq("WOS-3") & scoped["fst"].gt(0)
    w3 = metric_sum(scoped, w3_mask, "fst")
    w3_norm_source = scoped.assign(_w3_norm_lst=scoped["lst"].where(w3_mask & scoped["status"].eq("Normal"), 0.0))
    w3_norm_lst = metric_sum(w3_norm_source, w3_norm_source["_w3_norm_lst"].gt(0), "_w3_norm_lst")
    hi_w3_source = scoped.assign(_hi_w3=scoped["fst"].where(w3_mask & scoped["is_high"], 0.0))
    hi_w3 = metric_sum(hi_w3_source, hi_w3_source["_hi_w3"].gt(0), "_hi_w3")
    bsa_sales = allocate_bsa_scoped(bsa, booking, allowed)
    q2_progress = scoped.loc[scoped["fiscal_week"].isin(Q2_PROGRESS_WEEKS)].copy()
    q2_w3_mask = q2_progress["lead_time"].eq("WOS-3") & q2_progress["fst"].gt(0)
    q2_w3 = metric_sum(q2_progress, q2_w3_mask, "fst")
    q2_w3_norm_source = q2_progress.assign(
        _w3_norm_lst=q2_progress["lst"].where(q2_w3_mask & q2_progress["status"].eq("Normal"), 0.0)
    )
    q2_w3_norm_lst = metric_sum(q2_w3_norm_source, q2_w3_norm_source["_w3_norm_lst"].gt(0), "_w3_norm_lst")
    q2_hi_w3_source = q2_progress.assign(_hi_w3=q2_progress["fst"].where(q2_w3_mask & q2_progress["is_high"], 0.0))
    q2_hi_w3 = metric_sum(q2_hi_w3_source, q2_hi_w3_source["_hi_w3"].gt(0), "_hi_w3")
    q2_progress_bsa_sales = allocate_bsa_scoped(q2_progress_bsa, booking, allowed)
    return {
        "lst": lst,
        "w3": w3,
        "w3_norm_lst": w3_norm_lst,
        "hi_w3": hi_w3,
        "bsa": bsa_sales,
        "w3_q2_progress": q2_w3,
        "w3_norm_lst_q2_progress": q2_w3_norm_lst,
        "hi_w3_q2_progress": q2_hi_w3,
        "bsa_q2_progress": q2_progress_bsa_sales,
    }


def period_sum(frame: pd.DataFrame, origin: str, sales: str, months: set[str]) -> float:
    if frame.empty:
        return 0.0
    matched = frame.loc[frame["tab"].eq(origin) & frame["sales"].eq(sales) & frame["yyyymm"].isin(months), "value"]
    return float(matched.sum()) if not matched.empty else 0.0


def raw_for_sales(frames: dict[str, pd.DataFrame], origin: str, sales: str) -> dict[str, float]:
    return {
        "lst_2025": period_sum(frames["lst"], origin, sales, MONTHS_2025),
        "w3_2025": period_sum(frames["w3"], origin, sales, MONTHS_2025),
        "bsa_2025": period_sum(frames["bsa"], origin, sales, MONTHS_2025),
        "w3_norm_lst_2025": period_sum(frames["w3_norm_lst"], origin, sales, MONTHS_2025),
        "hi_w3_2025": period_sum(frames["hi_w3"], origin, sales, MONTHS_2025),
        "w3_q1": period_sum(frames["w3"], origin, sales, Q1_2026),
        "bsa_q1": period_sum(frames["bsa"], origin, sales, Q1_2026),
        "w3_norm_lst_q1": period_sum(frames["w3_norm_lst"], origin, sales, Q1_2026),
        "hi_w3_q1": period_sum(frames["hi_w3"], origin, sales, Q1_2026),
        "w3_q2_progress": period_sum(frames["w3_q2_progress"], origin, sales, Q2_2026),
        "bsa_q2_progress": period_sum(frames["bsa_q2_progress"], origin, sales, Q2_2026),
        "w3_norm_lst_q2_progress": period_sum(frames["w3_norm_lst_q2_progress"], origin, sales, Q2_2026),
        "hi_w3_q2_progress": period_sum(frames["hi_w3_q2_progress"], origin, sales, Q2_2026),
    }


def build_display_row(origin: str, sales: str, raw: dict[str, float], team_raw: dict[str, float], is_total: bool) -> dict[str, Any]:
    return {
        "tab": origin,
        "sales": "Team Total" if is_total else sales,
        "row_type": "TOTAL" if is_total else "SALES",
        "share_2025": 1.0 if is_total else safe_ratio(raw["lst_2025"], team_raw["lst_2025"]),
        "booking_base_2025": safe_ratio(raw["w3_2025"], raw["bsa_2025"]),
        "booking_q1_perform": safe_ratio(raw["w3_q1"], raw["bsa_q1"]),
        "booking_q2_progress": safe_ratio(raw["w3_q2_progress"], raw["bsa_q2_progress"]),
        "lifting_base_2025": safe_ratio(raw["w3_norm_lst_2025"], raw["w3_2025"]),
        "lifting_q1_perform": safe_ratio(raw["w3_norm_lst_q1"], raw["w3_q1"]),
        "lifting_q2_progress": safe_ratio(raw["w3_norm_lst_q2_progress"], raw["w3_q2_progress"]),
        "high_profit_base_2025": safe_ratio(raw["hi_w3_2025"], raw["w3_2025"]),
        "high_profit_q1_perform": safe_ratio(raw["hi_w3_q1"], raw["w3_q1"]),
        "high_profit_q2_progress": safe_ratio(raw["hi_w3_q2_progress"], raw["w3_q2_progress"]),
        "sort_lift_2025": raw["lst_2025"],
    }


def rows_for_origin(
    mod: Any,
    origin: str,
    entries: list[OwnerEntry],
    frames: dict[str, pd.DataFrame],
    explicit_order: bool,
) -> tuple[list[dict[str, Any]], dict[str, dict[str, float]]]:
    raw_by_sales = {
        entry.resolved_sales: raw_for_sales(frames, origin, entry.resolved_sales)
        for entry in entries
        if entry.resolved_sales and entry.match_status != "unmatched"
    }
    if not explicit_order:
        raw_by_sales = {
            sales: raw
            for sales, raw in raw_by_sales.items()
            if raw["lst_2025"] > 0 or raw["w3_2025"] > 0 or raw["w3_q1"] > 0 or raw["bsa_2025"] > 0 or raw["bsa_q1"] > 0
        }
    team_raw = {key: 0.0 for key in RAW_KEYS}
    for raw in raw_by_sales.values():
        for key in RAW_KEYS:
            team_raw[key] += raw[key]
    rows = [build_display_row(origin, "Team Total", team_raw, team_raw, is_total=True)]
    sales_order = list(raw_by_sales)
    if not explicit_order:
        sales_order.sort(key=lambda s: (-raw_by_sales[s]["lst_2025"], s))
    for sales in sales_order:
        rows.append(build_display_row(origin, sales, raw_by_sales[sales], team_raw, is_total=False))
    return rows, raw_by_sales


def account_counts(booking: pd.DataFrame, owner_entries: dict[str, list[OwnerEntry]]) -> dict[tuple[str, str], tuple[int, int, float | None]]:
    allowed = allowed_sales_by_origin(owner_entries)
    q1 = scoped_booking(booking, allowed)
    q1 = q1.loc[q1["yyyymm"].isin(Q1_2026) & q1["fst"].gt(0) & q1["shipper_code_key"].ne("")]
    out: dict[tuple[str, str], tuple[int, int, float | None]] = {}
    for origin, entries in owner_entries.items():
        sales_names = [entry.resolved_sales for entry in entries if entry.resolved_sales and entry.match_status != "unmatched"]
        team_shipper: set[str] = set()
        team_w3_shipper: set[str] = set()
        for sales in sales_names:
            part = q1.loc[q1["tab"].eq(origin) & q1["sales"].eq(sales)]
            total_set = set(part["shipper_code_key"].dropna().astype(str))
            w3_set = set(part.loc[part["lead_time"].eq("WOS-3"), "shipper_code_key"].dropna().astype(str))
            total = len(total_set)
            w3 = len(w3_set)
            out[(origin, sales)] = (total, w3, w3 / total if total else None)
            team_shipper |= total_set
            team_w3_shipper |= w3_set
        total = len(team_shipper)
        w3 = len(team_w3_shipper)
        out[(origin, "Team Total")] = (total, w3, w3 / total if total else None)
    return out


def blank_none(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return ""
    return value


def build_origin_values(mod: Any, origin: str, rows: list[dict[str, Any]], counts: dict[tuple[str, str], tuple[int, int, float | None]]) -> list[list[Any]]:
    values: list[list[Any]] = [[f"{origin} - 2026 OBT Sales Target & Performance"]]
    values.extend(mod.header_block(include_tab_col=False))
    target_cols = ("D", "G", "J", "M", "P", "S")
    perform_cols = ("E", "H", "K", "N", "Q", "T")
    base_cols = {"booking": "Y", "lifting": "Z", "hp": "AA"}
    for item in rows:
        sheet_row = len(values) + 1
        row_counts = counts.get((origin, item["sales"]), (0, 0, None))
        values.append(
            mod.make_data_row(
                item=item,
                sheet_row=sheet_row,
                counts=row_counts,
                base_cols=base_cols,
                origin_literal=origin,
                target_cols=target_cols,
                perform_cols=perform_cols,
                leading=[],
            )
        )
    return values


def build_summary_values(
    mod: Any,
    origins: list[str],
    rows_by_origin: dict[str, list[dict[str, Any]]],
    counts: dict[tuple[str, str], tuple[int, int, float | None]],
) -> list[list[Any]]:
    values: list[list[Any]] = [["2026 OBT Sales Target - All Origins"]]
    values.extend(mod.header_block(include_tab_col=True))
    target_cols = ("E", "H", "K", "N", "Q", "T")
    perform_cols = ("F", "I", "L", "O", "R", "U")
    base_cols = {"booking": "Z", "lifting": "AA", "hp": "AB"}
    for origin in origins:
        for item in rows_by_origin.get(origin, []):
            sheet_row = len(values) + 1
            row_counts = counts.get((origin, item["sales"]), (0, 0, None))
            values.append(
                mod.make_data_row(
                    item=item,
                    sheet_row=sheet_row,
                    counts=row_counts,
                    base_cols=base_cols,
                    origin_literal=None,
                    target_cols=target_cols,
                    perform_cols=perform_cols,
                    leading=[origin],
                )
            )
    return values


def compute_suggestions(mod: Any, rows_by_origin: dict[str, list[dict[str, Any]]]) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for origin, rows in rows_by_origin.items():
        if not rows:
            continue
        team = rows[0]
        metrics = {
            "booking": {"base": team["booking_base_2025"], "perf": team["booking_q1_perform"], "is_hp": False},
            "lifting": {"base": team["lifting_base_2025"], "perf": team["lifting_q1_perform"], "is_hp": False},
            "hp": {"base": team["high_profit_base_2025"], "perf": team["high_profit_q1_perform"], "is_hp": True},
        }
        out_metrics: dict[str, Any] = {}
        for name, metric in metrics.items():
            pp, why = mod.suggest_pp(metric["base"], metric["perf"], is_hp=metric["is_hp"])
            out_metrics[name] = {"base": metric["base"], "perf": metric["perf"], "pp": pp, "why": why}
        out[origin] = {
            "booking_pp": out_metrics["booking"]["pp"],
            "lifting_pp": out_metrics["lifting"]["pp"],
            "hp_pp": out_metrics["hp"]["pp"],
            "metrics": out_metrics,
        }
    return out


def fmt_pct(value: float | None) -> str:
    if value is None:
        return "n/a"
    return f"{value * 100:.1f}%"


def rationale(suggestion: dict[str, Any]) -> str:
    metrics = suggestion.get("metrics", {})
    labels = {"booking": "Booking", "lifting": "Lifting", "hp": "High-profit"}
    parts = []
    for key in ["booking", "lifting", "hp"]:
        item = metrics.get(key, {})
        pp = item.get("pp", 0.0)
        parts.append(f"{labels[key]} base {fmt_pct(item.get('base'))}, Q1 {fmt_pct(item.get('perf'))}, suggested +{pp * 100:.1f}%p")
    return "\n".join(parts)


def build_input_values(
    origins: list[str],
    existing: dict[str, tuple[float, float, float, str]],
    suggestions: dict[str, dict[str, Any]],
) -> list[list[Any]]:
    values: list[list[Any]] = [
        ["Origin-level 2026 target input and suggestions"],
        ["Input columns are percentage-point additions to the 2025 base. Per-origin tabs reference columns B, D, and F."],
        [
            "Tab",
            "3W Booking Input (%p)",
            "3W Booking Suggested (%p)",
            "Actual Lifting Input (%p)",
            "Actual Lifting Suggested (%p)",
            "High-Profit Input (%p)",
            "High-Profit Suggested (%p)",
            "Memo",
        ],
    ]
    for origin in origins:
        booking, lifting, hp, memo = existing.get(origin, (0.10, 0.10, 0.10, ""))
        suggestion = suggestions.get(origin, {})
        values.append(
            [
                origin,
                booking,
                suggestion.get("booking_pp", 0.10),
                lifting,
                suggestion.get("lifting_pp", 0.10),
                hp,
                suggestion.get("hp_pp", 0.10),
                rationale(suggestion) or memo,
            ]
        )
    return values


def build_readme_values() -> list[list[Any]]:
    return [
        ["2026 OBT Sales Target by Origin"],
        [""],
        ["This version recalculates salesperson performance from the current customer-code owner mapping."],
        ["Owner source", f"https://docs.google.com/spreadsheets/d/{SALES_OWNER_SOURCE_ID}/edit"],
        ["Customer owner mapping", "saleman.csv active rows as of 2026-05-19"],
        ["Performance basis", "2025, 2026 Q1, and 2026 Q2 week 14-19 bookings are attributed by BKG_SHPR_CST_NO -> current SALESMAN_NO."],
        ["BSA allocation", "Route BSA is allocated inside each tab's selected/current owner group by 2025 Normal LST_TEU route share; Q2 progress uses WW14-19 BSA."],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]


def build_owner_input_values(
    owner_entries: dict[str, list[OwnerEntry]],
    origins: list[str],
    raw_by_origin: dict[str, dict[str, dict[str, float]]],
) -> list[list[Any]]:
    values: list[list[Any]] = [
        ["Sales owner input"],
        [f"Generated at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} from current customer-code mapping"],
        [
            "Active",
            "Target_Tab",
            "Salesman",
            "Input_Name",
            "Match_Status",
            "Source_Type",
            "Source_Sheet",
            "Source_Cell",
            "Active_Customer_Count",
            "2025_LST_TEU",
            "2025_WOS3_FST_TEU",
            "2025_Allocated_BSA",
            "2026_Q1_WOS3_FST_TEU",
            "2026_Q1_Allocated_BSA",
            "Memo",
        ],
    ]
    for origin in origins:
        entries = owner_entries.get(origin, [])
        if not entries:
            values.append(["", origin, "", "", "no_current_owner", "", "", "", 0, 0, 0, 0, 0, 0, "No active owner/customer performance found for this tab."])
            continue
        for entry in entries:
            raw = raw_by_origin.get(origin, {}).get(entry.resolved_sales, {k: 0.0 for k in RAW_KEYS})
            memo = "" if entry.match_status != "unmatched" else "Name from Contact Point was not found in active salesperson CSV SALESMAN_NO values."
            values.append(
                [
                    "Y" if entry.match_status != "unmatched" else "N",
                    origin,
                    entry.resolved_sales,
                    entry.input_name,
                    entry.match_status,
                    entry.source_type,
                    entry.source_sheet,
                    entry.source_cell,
                    entry.customer_count,
                    raw["lst_2025"],
                    raw["w3_2025"],
                    raw["bsa_2025"],
                    raw["w3_q1"],
                    raw["bsa_q1"],
                    memo,
                ]
            )
    return values


def rgb(hex_color: str) -> dict[str, float]:
    clean = hex_color.strip().lstrip("#")
    return {
        "red": int(clean[0:2], 16) / 255,
        "green": int(clean[2:4], 16) / 255,
        "blue": int(clean[4:6], 16) / 255,
    }


def owner_input_format_requests(sheet_id: int, row_count: int) -> list[dict[str, Any]]:
    last_data_row = max(row_count, 4)
    border = {"style": "SOLID", "color": rgb("B7B7B7")}
    return [
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": sheet_id,
                    "gridProperties": {
                        "rowCount": last_data_row + 20,
                        "columnCount": 15,
                        "frozenRowCount": 3,
                    },
                },
                "fields": "gridProperties(rowCount,columnCount,frozenRowCount)",
            }
        },
        {
            "mergeCells": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": 15,
                },
                "mergeType": "MERGE_ALL",
            }
        },
        {
            "mergeCells": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": 2,
                    "startColumnIndex": 0,
                    "endColumnIndex": 15,
                },
                "mergeType": "MERGE_ALL",
            }
        },
        {
            "repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1, "startColumnIndex": 0, "endColumnIndex": 15},
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": rgb("D9EAD3"),
                        "textFormat": {"bold": True, "fontSize": 12},
                        "horizontalAlignment": "LEFT",
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)",
            }
        },
        {
            "repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": 2, "endRowIndex": 3, "startColumnIndex": 0, "endColumnIndex": 15},
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": rgb("CFE2F3"),
                        "textFormat": {"bold": True},
                        "horizontalAlignment": "CENTER",
                        "verticalAlignment": "MIDDLE",
                        "wrapStrategy": "WRAP",
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment,wrapStrategy)",
            }
        },
        {
            "repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": 3, "endRowIndex": last_data_row, "startColumnIndex": 0, "endColumnIndex": 15},
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "CENTER",
                        "verticalAlignment": "MIDDLE",
                        "wrapStrategy": "WRAP",
                    }
                },
                "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,wrapStrategy)",
            }
        },
        {
            "repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": 3, "endRowIndex": last_data_row, "startColumnIndex": 8, "endColumnIndex": 14},
                "cell": {"userEnteredFormat": {"numberFormat": {"type": "NUMBER", "pattern": "#,##0.0"}}},
                "fields": "userEnteredFormat.numberFormat",
            }
        },
        {
            "updateBorders": {
                "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": last_data_row, "startColumnIndex": 0, "endColumnIndex": 15},
                "top": border,
                "bottom": border,
                "left": border,
                "right": border,
                "innerHorizontal": border,
                "innerVertical": border,
            }
        },
        {
            "addConditionalFormatRule": {
                "rule": {
                    "ranges": [
                        {"sheetId": sheet_id, "startRowIndex": 3, "endRowIndex": last_data_row, "startColumnIndex": 0, "endColumnIndex": 15}
                    ],
                    "booleanRule": {
                        "condition": {"type": "CUSTOM_FORMULA", "values": [{"userEnteredValue": '=$A4="N"'}]},
                        "format": {"backgroundColor": rgb("FCE4D6"), "textFormat": {"foregroundColor": rgb("990000")}},
                    },
                },
                "index": 0,
            }
        },
        {
            "updateDimensionProperties": {
                "range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": 0, "endIndex": 1},
                "properties": {"pixelSize": 70},
                "fields": "pixelSize",
            }
        },
        {
            "updateDimensionProperties": {
                "range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": 1, "endIndex": 8},
                "properties": {"pixelSize": 125},
                "fields": "pixelSize",
            }
        },
        {
            "updateDimensionProperties": {
                "range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": 8, "endIndex": 14},
                "properties": {"pixelSize": 115},
                "fields": "pixelSize",
            }
        },
        {
            "updateDimensionProperties": {
                "range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": 14, "endIndex": 15},
                "properties": {"pixelSize": 260},
                "fields": "pixelSize",
            }
        },
    ]


def reset_format_requests(sheet_id: int, conditional_count: int) -> list[dict[str, Any]]:
    requests: list[dict[str, Any]] = []
    for index in range(conditional_count - 1, -1, -1):
        requests.append({"deleteConditionalFormatRule": {"sheetId": sheet_id, "index": index}})
    requests.append(
        {
            "unmergeCells": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": 2000,
                    "startColumnIndex": 0,
                    "endColumnIndex": 30,
                }
            }
        }
    )
    requests.append(
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": 2000,
                    "startColumnIndex": 0,
                    "endColumnIndex": 30,
                },
                "cell": {"userEnteredFormat": {}},
                "fields": "userEnteredFormat",
            }
        }
    )
    return requests


def apply_target_formatting(
    service: Any,
    spreadsheet_id: str,
    mod: Any,
    origins: list[str],
    origin_rowcounts: dict[str, int],
    rowcounts: dict[str, int],
) -> None:
    meta = with_backoff(
        service.spreadsheets().get(
            spreadsheetId=spreadsheet_id,
            fields="sheets(properties(sheetId,title),conditionalFormats)",
        )
    )
    sheet_meta = {sheet["properties"]["title"]: sheet for sheet in meta.get("sheets", [])}
    sheet_ids = {title: sheet["properties"]["sheetId"] for title, sheet in sheet_meta.items()}
    required = [mod.README_SHEET, mod.INPUT_SHEET, SALES_OWNER_INPUT_SHEET, mod.SUMMARY_SHEET, *origins]
    requests: list[dict[str, Any]] = []
    for title in required:
        if title not in sheet_ids:
            continue
        requests.extend(reset_format_requests(sheet_ids[title], len(sheet_meta[title].get("conditionalFormats", []))))
    requests.extend(mod.readme_format_requests(sheet_ids[mod.README_SHEET], rowcounts[mod.README_SHEET]))
    requests.extend(mod.input_format_requests(sheet_ids[mod.INPUT_SHEET], rowcounts[mod.INPUT_SHEET]))
    requests.extend(owner_input_format_requests(sheet_ids[SALES_OWNER_INPUT_SHEET], rowcounts[SALES_OWNER_INPUT_SHEET]))
    requests.extend(mod.summary_format_requests(sheet_ids[mod.SUMMARY_SHEET], rowcounts[mod.SUMMARY_SHEET]))
    for origin in origins:
        requests.extend(mod.origin_format_requests(sheet_ids[origin], origin_rowcounts[origin]))
    mod.batch_apply(service, spreadsheet_id, requests)


def ensure_support_sheet(service: Any, spreadsheet_id: str) -> dict[str, int]:
    meta = with_backoff(service.spreadsheets().get(spreadsheetId=spreadsheet_id, fields="sheets(properties(sheetId,title,index,gridProperties))"))
    sheet_ids = {s["properties"]["title"]: s["properties"]["sheetId"] for s in meta.get("sheets", [])}
    if SALES_OWNER_INPUT_SHEET not in sheet_ids:
        with_backoff(
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={
                    "requests": [
                        {
                            "addSheet": {
                                "properties": {
                                    "title": SALES_OWNER_INPUT_SHEET,
                                    "index": 3,
                                    "gridProperties": {"rowCount": 500, "columnCount": 16},
                                }
                            }
                        }
                    ]
                },
            )
        )
        meta = with_backoff(service.spreadsheets().get(spreadsheetId=spreadsheet_id, fields="sheets(properties(sheetId,title,index,gridProperties))"))
        sheet_ids = {s["properties"]["title"]: s["properties"]["sheetId"] for s in meta.get("sheets", [])}
    else:
        props = next(s["properties"] for s in meta.get("sheets", []) if s["properties"]["title"] == SALES_OWNER_INPUT_SHEET)
        grid = props.get("gridProperties", {})
        if grid.get("columnCount", 0) < 16 or grid.get("rowCount", 0) < 500:
            with_backoff(
                service.spreadsheets().batchUpdate(
                    spreadsheetId=spreadsheet_id,
                    body={
                        "requests": [
                            {
                                "updateSheetProperties": {
                                    "properties": {
                                        "sheetId": sheet_ids[SALES_OWNER_INPUT_SHEET],
                                        "gridProperties": {
                                            "rowCount": max(500, grid.get("rowCount", 0)),
                                            "columnCount": max(16, grid.get("columnCount", 0)),
                                        },
                                    },
                                    "fields": "gridProperties(rowCount,columnCount)",
                                }
                            }
                        ]
                    },
                )
            )
    return sheet_ids


def audit_mapping(booking: pd.DataFrame, owner_entries: dict[str, list[OwnerEntry]], raw_by_origin: dict[str, dict[str, dict[str, float]]]) -> pd.DataFrame:
    active_tabs = set(owner_entries)
    total_rows = int(len(booking))
    mapped_rows = int(booking["mapped_current_owner"].sum())
    rows = []
    for origin, entries in owner_entries.items():
        part = booking.loc[booking["tab"].eq(origin)]
        scoped_sales = {e.resolved_sales for e in entries if e.match_status != "unmatched"}
        scoped = part.loc[part["sales"].isin(scoped_sales)]
        rows.append(
            {
                "Target_Tab": origin,
                "Owner_Source": (
                    owner_entries.get(origin, [OwnerEntry(origin, "", "", "", "", "", "", "", 0)])[0].source_type
                    if owner_entries.get(origin)
                    else "fallback_active_mapping"
                ),
                "Owner_Count": len(entries),
                "Mapped_Booking_Rows": int(scoped["mapped_current_owner"].sum()),
                "Mapped_Customer_Count": int(scoped.loc[scoped["mapped_current_owner"], "shipper_code_key"].nunique()),
                "2025_LST_TEU": sum(raw.get("lst_2025", 0.0) for raw in raw_by_origin.get(origin, {}).values()),
                "2025_WOS3_FST_TEU": sum(raw.get("w3_2025", 0.0) for raw in raw_by_origin.get(origin, {}).values()),
                "2025_Allocated_BSA": sum(raw.get("bsa_2025", 0.0) for raw in raw_by_origin.get(origin, {}).values()),
                "2026_Q1_WOS3_FST_TEU": sum(raw.get("w3_q1", 0.0) for raw in raw_by_origin.get(origin, {}).values()),
                "2026_Q1_Allocated_BSA": sum(raw.get("bsa_q1", 0.0) for raw in raw_by_origin.get(origin, {}).values()),
                "2026_Q2_W14_19_WOS3_FST_TEU": sum(
                    raw.get("w3_q2_progress", 0.0) for raw in raw_by_origin.get(origin, {}).values()
                ),
                "2026_Q2_W14_19_Allocated_BSA": sum(
                    raw.get("bsa_q2_progress", 0.0) for raw in raw_by_origin.get(origin, {}).values()
                ),
            }
        )
    summary = pd.DataFrame(rows)
    summary.attrs["total_booking_rows"] = total_rows
    summary.attrs["mapped_booking_rows"] = mapped_rows
    summary.attrs["active_tabs"] = len(active_tabs)
    return summary


def write_audit_files(report: dict[str, Any], owner_entries: dict[str, list[OwnerEntry]], audit: pd.DataFrame) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    owner_rows = []
    for entries in owner_entries.values():
        for entry in entries:
            owner_rows.append(entry.__dict__)
    pd.DataFrame(owner_rows).to_csv(OUT_DIR / "current_customer_owner_input_audit_20260519.csv", index=False, encoding="utf-8-sig")
    audit.to_csv(OUT_DIR / "current_customer_owner_target_audit_20260519.csv", index=False, encoding="utf-8-sig")
    (OUT_DIR / "current_customer_owner_target_update_20260519.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Build payloads and audit files without writing to Google Sheets.")
    args = parser.parse_args()

    mod = target_builder()
    creds = get_creds()
    drive = build("drive", "v3", credentials=creds)
    sheets = build("sheets", "v4", credentials=creds)

    owner_path = export_sheet(drive, SALES_OWNER_SOURCE_ID, "contact_point_sales_owner_source_export_20260519.xlsx")
    target_path = export_sheet(drive, TARGET_SPREADSHEET_ID, "target_workbook_before_current_customer_owner_update_20260519.xlsx")

    active_sales, customer_to_sales, customer_counts = active_salesman_mapping()
    explicit_entries = load_owner_entries(owner_path, active_sales, customer_counts)
    booking = load_booking(customer_to_sales)
    bsa = load_bsa()
    q2_progress_bsa = load_bsa(Q2_PROGRESS_WEEKS)

    origins = target_origins_from_workbook(target_path, mod)
    fallback_sales = candidate_sales_by_origin(booking)
    owner_entries = complete_owner_entries(origins, explicit_entries, fallback_sales, customer_counts)
    allowed = allowed_sales_by_origin(owner_entries)
    frames = build_metric_frames(booking, bsa, q2_progress_bsa, allowed)
    counts = account_counts(booking, owner_entries)

    rows_by_origin: dict[str, list[dict[str, Any]]] = {}
    raw_by_origin: dict[str, dict[str, dict[str, float]]] = {}
    for origin in origins:
        explicit_order = bool(explicit_entries.get(origin))
        rows, raw = rows_for_origin(mod, origin, owner_entries.get(origin, []), frames, explicit_order)
        rows_by_origin[origin] = rows
        raw_by_origin[origin] = raw

    existing_input = read_existing_target_input(target_path, mod)
    suggestions = compute_suggestions(mod, rows_by_origin)
    summary_values = build_summary_values(mod, origins, rows_by_origin, counts)
    readme_values = build_readme_values()
    input_values = build_input_values(origins, existing_input, suggestions)
    owner_input_values = build_owner_input_values(owner_entries, origins, raw_by_origin)
    payloads: list[tuple[str, list[list[Any]]]] = [
        (mod.README_SHEET, readme_values),
        (mod.INPUT_SHEET, input_values),
        (SALES_OWNER_INPUT_SHEET, owner_input_values),
        (mod.SUMMARY_SHEET, summary_values),
    ]
    origin_rowcounts: dict[str, int] = {}
    for origin in origins:
        values = build_origin_values(mod, origin, rows_by_origin[origin], counts)
        payloads.append((origin, values))
        origin_rowcounts[origin] = len(values)

    audit = audit_mapping(booking, owner_entries, raw_by_origin)
    owner_row_count = sum(len(entries) for entries in owner_entries.values())
    unmatched = sum(1 for entries in owner_entries.values() for entry in entries if entry.match_status == "unmatched")
    report = {
        "target_spreadsheet_id": TARGET_SPREADSHEET_ID,
        "sales_owner_source_id": SALES_OWNER_SOURCE_ID,
        "customer_owner_file": str(ROOT / ("saleman.csv" if (ROOT / "saleman.csv").exists() else "salesman.csv")),
        "booking_rows": int(len(booking)),
        "q2_progress_weeks": "14-19",
        "booking_rows_mapped_to_current_owner": int(booking["mapped_current_owner"].sum()),
        "booking_customer_codes": int(booking["shipper_code_key"].nunique()),
        "mapped_customer_codes": int(booking.loc[booking["mapped_current_owner"], "shipper_code_key"].nunique()),
        "origins": len(origins),
        "owner_rows": owner_row_count,
        "unmatched_owner_rows": unmatched,
        "origin_rowcounts": origin_rowcounts,
        "dry_run": args.dry_run,
    }
    write_audit_files(report, owner_entries, audit)
    print(json.dumps(report, ensure_ascii=False, indent=2))

    if args.dry_run:
        return

    ensure_support_sheet(sheets, TARGET_SPREADSHEET_ID)
    mod.clear_tab_values(
        sheets,
        TARGET_SPREADSHEET_ID,
        [mod.README_SHEET, mod.INPUT_SHEET, SALES_OWNER_INPUT_SHEET, mod.SUMMARY_SHEET, *origins],
    )
    mod.batch_write_values(sheets, TARGET_SPREADSHEET_ID, payloads)
    apply_target_formatting(
        sheets,
        TARGET_SPREADSHEET_ID,
        mod,
        origins,
        origin_rowcounts,
        {
            mod.README_SHEET: len(readme_values),
            mod.INPUT_SHEET: len(input_values),
            SALES_OWNER_INPUT_SHEET: len(owner_input_values),
            mod.SUMMARY_SHEET: len(summary_values),
        },
    )


if __name__ == "__main__":
    main()
