from __future__ import annotations

import json
import os
import re
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "output"
YEAR = "2025"
BSA_TEAMS = ("OBT", "EST", "IST", "JBT")

FISCAL_445 = {
    2025: (datetime(2024, 12, 29), [4, 4, 5, 4, 4, 5, 4, 4, 5, 4, 4, 6]),
    2026: (datetime(2026, 1, 4), [4, 4, 5, 4, 4, 5, 4, 4, 5, 4, 4, 5]),
    2027: (datetime(2027, 1, 3), [4, 4, 5, 4, 4, 5, 4, 4, 5, 4, 4, 5]),
}


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [re.sub(r"[^\x00-\x7F]+$", "", c).strip() for c in df.columns]
    return df


def read_view1(path: Path) -> pd.DataFrame:
    usecols = [
        "BKG_NO",
        "BKG_SHPR_CST_ENM",
        "BKG_SHPR_CST_NO",
        "DLY_CTR_CD",
        "DLY_PLC_CD",
        "POR_CTR_CD",
        "POR_PLC_CD",
        "Booking_date일",
        "Booking_schedule일",
        "Cancel_date일",
        "FST_TEU",
    ]
    df = pd.read_csv(path, encoding="utf-8-sig", dtype=str, usecols=usecols)
    return clean_columns(df)


def read_view2(path: Path) -> pd.DataFrame:
    usecols = [
        "BKG_NO",
        "POR_Country",
        "POR_PORT",
        "DLY_Country",
        "DLY_PORT",
        "Date_vsl일",
        "Booking_status",
        "Salesman_POR",
        "CM1_Booking",
        "TEU_Booking",
    ]
    df = pd.read_csv(path, encoding="utf-16", sep="\t", dtype=str, usecols=usecols)
    return clean_columns(df)


def drop_total_rows(df: pd.DataFrame) -> pd.DataFrame:
    total_markers = {"전체", "Total", "Grand Total"}
    mask = pd.Series(False, index=df.index)
    for col in ["BKG_NO", "Booking_status", "LST_Status"]:
        if col in df.columns:
            mask = mask | df[col].astype(str).str.strip().isin(total_markers)
    return df.loc[~mask].copy() if mask.any() else df


def parse_korean_date(series: pd.Series) -> pd.Series:
    text = series.fillna("").astype(str)
    parts = text.str.extract(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})")
    parts = parts.apply(pd.to_numeric, errors="coerce")
    return pd.to_datetime(
        {"year": parts[0], "month": parts[1], "day": parts[2]},
        errors="coerce",
    )


def sunday_start(dates: pd.Series) -> pd.Series:
    offset = (dates.dt.dayofweek + 1) % 7
    return dates - pd.to_timedelta(offset, unit="D")


def build_445_map() -> dict[str, str]:
    out: dict[str, str] = {}
    for year, (first_sun, pattern) in FISCAL_445.items():
        wk = 0
        for mi, cnt in enumerate(pattern, start=1):
            ym = f"{year}{mi:02d}"
            for _ in range(cnt):
                out[(first_sun + timedelta(weeks=wk)).strftime("%Y-%m-%d")] = ym
                wk += 1
    return out


def classify_team(origin: pd.Series, dest: pd.Series) -> pd.Series:
    o = origin.fillna("").astype(str).str.strip()
    d = dest.fillna("").astype(str).str.strip()
    return pd.Series(
        np.select(
            [
                (~o.isin(["KR", "JP"])) & (d != "KR"),
                (o == "KR") & (d != "JP"),
                (o != "JP") & (d == "KR"),
            ],
            ["OBT", "EST", "IST"],
            default="JBT",
        ),
        index=origin.index,
    )


def to_number(series: pd.Series) -> pd.Series:
    return pd.to_numeric(
        series.fillna("").astype(str).str.replace(",", "", regex=False).str.strip(),
        errors="coerce",
    ).fillna(0.0)


def load_grade() -> pd.DataFrame:
    path = OUT_DIR / "grade_2025.csv"
    grade = pd.read_csv(path, encoding="utf-8-sig", comment="#", dtype=str)
    code_col = next(c for c in grade.columns if "Booking" in c or "Shipper" in c)
    grade_col = next((c for c in grade.columns if "grade" in c.lower()), grade.columns[-1])
    grade = grade[[code_col, grade_col]].rename(
        columns={code_col: "BKG_SHPR_CST_NO", grade_col: "grade"}
    )
    grade["BKG_SHPR_CST_NO"] = grade["BKG_SHPR_CST_NO"].fillna("").astype(str).str.strip()
    grade["grade"] = grade["grade"].fillna("").astype(str).str.strip().replace(
        {"AB": "A+B", "CD": "C+D"}
    )
    return grade.loc[grade["BKG_SHPR_CST_NO"].ne("")].drop_duplicates("BKG_SHPR_CST_NO")


def load_previous_actual_schedule() -> pd.Series:
    candidates = []
    for path in OUT_DIR.glob("booking_snapshot_result_*.csv"):
        m = re.search(r"(\d{8})", path.stem)
        if m:
            candidates.append((m.group(1), path.stat().st_mtime, path))
    for _, _, path in sorted(candidates, reverse=True):
        try:
            prev = pd.read_csv(
                path,
                dtype=str,
                encoding="utf-8-sig",
                usecols=["BKG_NO", "Actual_Departure_schedule"],
            )
            prev["BKG_NO"] = prev["BKG_NO"].fillna("").astype(str).str.strip()
            prev["Actual_Departure_schedule"] = (
                prev["Actual_Departure_schedule"].fillna("").astype(str).str.strip()
            )
            prev = prev.loc[prev["BKG_NO"].ne("") & prev["Actual_Departure_schedule"].ne("")]
            return prev.drop_duplicates("BKG_NO", keep="last").set_index("BKG_NO")[
                "Actual_Departure_schedule"
            ]
        except Exception:
            continue
    return pd.Series(dtype=str)


def build_booking_frame() -> pd.DataFrame:
    print("reading 2025 booking source files")
    df1 = drop_total_rows(read_view1(ROOT / "1_2025.csv"))
    df2 = drop_total_rows(read_view2(ROOT / "2_2025.csv"))

    df1 = df1.drop_duplicates("BKG_NO", keep="first")
    df2 = df2.drop_duplicates("BKG_NO", keep="first")

    df2 = df2.rename(
        columns={
            "POR_Country": "POR_CTR_CD",
            "POR_PORT": "POR_PLC_CD",
            "DLY_Country": "DLY_CTR_CD",
            "DLY_PORT": "DLY_PLC_CD",
            "Booking_status": "LST_Status",
            "CM1_Booking": "CM1",
            "TEU_Booking": "LST_TEU",
            "Date_vsl": "Actual_Departure_schedule",
        }
    )

    print(f"view1 rows={len(df1):,}, view2 unique rows={len(df2):,}")
    out = df2.merge(
        df1[
            [
                "BKG_NO",
                "BKG_SHPR_CST_NO",
                "BKG_SHPR_CST_ENM",
                "POR_CTR_CD",
                "POR_PLC_CD",
                "DLY_CTR_CD",
                "DLY_PLC_CD",
                "Booking_date",
                "Booking_schedule",
                "Cancel_date",
                "FST_TEU",
            ]
        ],
        on="BKG_NO",
        how="left",
        suffixes=("", "_v1"),
    )

    for col in ["POR_CTR_CD", "POR_PLC_CD", "DLY_CTR_CD", "DLY_PLC_CD"]:
        fallback = f"{col}_v1"
        if fallback in out.columns:
            empty = out[col].fillna("").astype(str).str.strip().isin(["", "nan", "None"])
            out.loc[empty, col] = out.loc[empty, fallback]
            out = out.drop(columns=[fallback])

    for col in ["Booking_date", "Booking_schedule"]:
        empty = out[col].fillna("").astype(str).str.strip().isin(["", "nan", "None"])
        out.loc[empty, col] = out.loc[empty, "Actual_Departure_schedule"]

    out["FST_TEU"] = out["FST_TEU"].where(
        ~out["FST_TEU"].fillna("").astype(str).str.strip().isin(["", "nan", "None"]),
        out["LST_TEU"],
    )

    missing = df1.loc[~df1["BKG_NO"].isin(set(df2["BKG_NO"]))].copy()
    cancel_key = missing["Cancel_date"].fillna("").astype(str).str.strip()
    cancel_missing = missing.loc[cancel_key.ne("") & ~cancel_key.str.lower().isin(["nan", "none", "nat"])].copy()
    prev_actual = load_previous_actual_schedule()
    if not cancel_missing.empty and not prev_actual.empty:
        actual = cancel_missing["BKG_NO"].astype(str).str.strip().map(prev_actual).fillna("")
        cancel_missing = cancel_missing.loc[actual.ne("")].copy()
        actual = actual.loc[actual.ne("")]
        recovered = pd.DataFrame(
            {
                "BKG_NO": cancel_missing["BKG_NO"].values,
                "POR_CTR_CD": cancel_missing["POR_CTR_CD"].values,
                "POR_PLC_CD": cancel_missing["POR_PLC_CD"].values,
                "DLY_CTR_CD": cancel_missing["DLY_CTR_CD"].values,
                "DLY_PLC_CD": cancel_missing["DLY_PLC_CD"].values,
                "Actual_Departure_schedule": actual.values,
                "LST_Status": "Cancel",
                "Salesman_POR": "",
                "CM1": "",
                "LST_TEU": "0",
                "BKG_SHPR_CST_NO": cancel_missing["BKG_SHPR_CST_NO"].values,
                "BKG_SHPR_CST_ENM": cancel_missing["BKG_SHPR_CST_ENM"].values,
                "Booking_date": cancel_missing["Booking_date"].values,
                "Booking_schedule": cancel_missing["Booking_schedule"].values,
                "Cancel_date": cancel_missing["Cancel_date"].values,
                "FST_TEU": cancel_missing["FST_TEU"].values,
            }
        )
        out = pd.concat([out, recovered[out.columns]], ignore_index=True)
        print(f"recovered cancel rows={len(recovered):,}")

    return out


def add_dashboard_fields(df: pd.DataFrame) -> pd.DataFrame:
    print("computing dashboard fields")
    df = df.copy()
    for col in [
        "BKG_SHPR_CST_NO",
        "BKG_SHPR_CST_ENM",
        "POR_CTR_CD",
        "POR_PLC_CD",
        "DLY_CTR_CD",
        "DLY_PLC_CD",
        "LST_Status",
        "Salesman_POR",
    ]:
        df[col] = df[col].fillna("").astype(str).str.strip()

    date_booking = parse_korean_date(df["Booking_date"])
    date_schedule = parse_korean_date(df["Booking_schedule"])
    date_cancel = parse_korean_date(df["Cancel_date"])
    date_actual = parse_korean_date(df["Actual_Departure_schedule"])

    week_start_actual = sunday_start(date_actual)
    week_start_schedule = sunday_start(date_schedule)
    df["week_start_date_key"] = week_start_actual.dt.strftime("%Y-%m-%d").fillna("")
    df["week_start_bkg_sche_key"] = week_start_schedule.dt.strftime("%Y-%m-%d").fillna("")
    df["YYYYMM"] = df["week_start_date_key"].map(build_445_map()).fillna("")

    diff_bkg_sche = (week_start_schedule - date_booking).dt.days
    df["Lead_time_BKG_Sche"] = np.select(
        [
            diff_bkg_sche < 1,
            diff_bkg_sche <= 7,
            diff_bkg_sche <= 14,
            diff_bkg_sche.notna(),
        ],
        ["Week of Sailing (WOS)", "WOS-1", "WOS-2", "WOS-3"],
        default="",
    )

    df["fst"] = to_number(df["FST_TEU"])
    df["lst"] = to_number(df["LST_TEU"])
    df["cm1v"] = to_number(df["CM1"])
    df = df.merge(load_grade(), on="BKG_SHPR_CST_NO", how="left")
    df["grade"] = df["grade"].fillna("").astype(str).str.strip()

    status = df["LST_Status"]
    is_cancel = status.eq("Cancel")
    diff_pn = (date_cancel - date_booking).dt.days
    diff_rn = (date_actual - date_booking).dt.days
    exclude = (
        is_cancel & date_cancel.notna() & date_booking.notna() & (diff_pn <= 3)
    ) | (
        is_cancel
        & date_actual.notna()
        & date_booking.notna()
        & date_cancel.notna()
        & (diff_rn >= 21)
        & (diff_pn <= 7)
    )
    before = len(df)
    df = df.loc[status.ne("") & ~exclude].copy()
    df = df.loc[df["YYYYMM"].astype(str).str.startswith(YEAR)].copy()
    print(f"filtered rows={len(df):,} (removed {before - len(df):,})")

    df["team"] = classify_team(df["POR_CTR_CD"], df["DLY_CTR_CD"])
    df["is_normal"] = df["LST_Status"].eq("Normal")
    df["is_w3"] = df["Lead_time_BKG_Sche"].eq("WOS-3")

    # 루트별 고/저 (POR_PLC_CD+DLY_PLC_CD 루트 평균 CM1/TEU 대비 화주 CM1/TEU).
    # daily_3w_dashboard / build_salesperson_bsa_action_sheet 와 동일 정의.
    # 목표 파이프라인(update_target_workbook)이 2025 base를 이 컬럼으로 계산하도록 캐시에 포함.
    print("computing route 고/저 (루트 CM1/TEU 평균 대비)")
    rmask = df["is_normal"] & df["cm1v"].ne(0) & df["lst"].gt(0)
    rvalid = df.loc[rmask, ["BKG_SHPR_CST_NO", "POR_PLC_CD", "DLY_PLC_CD", "cm1v", "lst"]].copy()
    route_agg = rvalid.groupby(["POR_PLC_CD", "DLY_PLC_CD"], dropna=False).agg(
        r_cm1=("cm1v", "sum"), r_teu=("lst", "sum")
    )
    route_agg["r_avg"] = route_agg["r_cm1"] / route_agg["r_teu"]
    shpr_route = rvalid.groupby(["BKG_SHPR_CST_NO", "POR_PLC_CD", "DLY_PLC_CD"], dropna=False).agg(
        s_cm1=("cm1v", "sum"), s_teu=("lst", "sum")
    )
    shpr_route["s_avg"] = shpr_route["s_cm1"] / shpr_route["s_teu"]
    shpr_route = shpr_route.join(route_agg[["r_avg"]], on=["POR_PLC_CD", "DLY_PLC_CD"])
    shpr_route["pt"] = np.where(shpr_route["s_avg"] >= shpr_route["r_avg"], "고수익", "저수익")
    _pt_lookup = shpr_route["pt"].to_dict()
    df["고/저"] = [
        _pt_lookup.get((s, p, d), "")
        for s, p, d in zip(df["BKG_SHPR_CST_NO"], df["POR_PLC_CD"], df["DLY_PLC_CD"])
    ]

    print("computing high-shipper tag")
    valid = df.loc[df["is_normal"] & df["cm1v"].ne(0) & df["lst"].gt(0)].copy()
    valid = valid.loc[valid["YYYYMM"].ne("")]
    por_month = valid.groupby(["POR_PLC_CD", "YYYYMM"], dropna=False).agg(
        p_cm1=("cm1v", "sum"), p_teu=("lst", "sum")
    )
    por_month["p_avg"] = por_month["p_cm1"] / por_month["p_teu"]

    shpr_por = valid.groupby(["BKG_SHPR_CST_NO", "POR_PLC_CD", "YYYYMM"], dropna=False).agg(
        s_cm1=("cm1v", "sum"), s_teu=("lst", "sum")
    )
    shpr_por["s_avg"] = shpr_por["s_cm1"] / shpr_por["s_teu"]
    shpr_por = shpr_por.join(por_month[["p_avg"]], on=["POR_PLC_CD", "YYYYMM"])
    shpr_por["tag"] = np.where(shpr_por["s_avg"] >= shpr_por["p_avg"], "고수익화주", "저수익화주")
    tag_por = shpr_por["tag"].to_dict()

    all_month = valid.groupby(["YYYYMM"], dropna=False).agg(a_cm1=("cm1v", "sum"), a_teu=("lst", "sum"))
    all_month["a_avg"] = all_month["a_cm1"] / all_month["a_teu"]
    shpr_all = valid.groupby(["BKG_SHPR_CST_NO", "YYYYMM"], dropna=False).agg(
        s_cm1=("cm1v", "sum"), s_teu=("lst", "sum")
    )
    shpr_all["s_avg"] = shpr_all["s_cm1"] / shpr_all["s_teu"]
    shpr_all = shpr_all.join(all_month[["a_avg"]], on=["YYYYMM"])
    shpr_all["tag"] = np.where(shpr_all["s_avg"] >= shpr_all["a_avg"], "고수익화주", "저수익화주")
    tag_all = shpr_all["tag"].to_dict()

    pairs = df[["BKG_SHPR_CST_NO", "POR_PLC_CD", "grade", "YYYYMM"]].copy()
    latest = pairs.loc[pairs["YYYYMM"].ne("")].groupby(
        ["BKG_SHPR_CST_NO", "POR_PLC_CD"], dropna=False
    )["YYYYMM"].max()
    pairs = pairs.drop_duplicates(["BKG_SHPR_CST_NO", "POR_PLC_CD"]).set_index(
        ["BKG_SHPR_CST_NO", "POR_PLC_CD"]
    )
    pairs["latest_YYYYMM"] = latest

    def choose_tag(row: pd.Series) -> str:
        shipper, por = row.name
        cur_month = row.get("latest_YYYYMM", "")
        if not cur_month or pd.isna(cur_month) or not shipper:
            return ""
        y = int(str(cur_month)[:4])
        m = int(str(cur_month)[4:])
        for i in range(1, 7):
            pm = m - i
            py = y
            while pm <= 0:
                pm += 12
                py -= 1
            ym = f"{py}{pm:02d}"
            tag = tag_por.get((shipper, por, ym))
            if tag:
                return tag
            tag = tag_all.get((shipper, ym))
            if tag:
                return tag
        tag = tag_por.get((shipper, por, str(cur_month)))
        if tag:
            return tag
        grade = str(row.get("grade", "")).strip()
        if grade == "C+D":
            return "고수익화주"
        if grade == "A+B":
            return "저수익화주"
        return ""

    pair_tag = pairs.apply(choose_tag, axis=1).to_dict()
    df["고수익화주 태그"] = [
        pair_tag.get((shipper, por), "")
        for shipper, por in zip(df["BKG_SHPR_CST_NO"], df["POR_PLC_CD"])
    ]
    return df


def load_bsa() -> pd.DataFrame:
    path = sorted(OUT_DIR.glob("BSA_raw_monthly3W_*.csv"), key=os.path.getmtime, reverse=True)[0]
    bsa = pd.read_csv(path, dtype=str, encoding="utf-8-sig")
    bsa["team"] = bsa["Sales Team"].fillna("").astype(str).str.strip().str.upper()
    bsa = bsa.loc[bsa["team"].isin(BSA_TEAMS)].copy()
    bsa = bsa.loc[bsa["YYYYMM"].fillna("").astype(str).str.startswith(YEAR)].copy()
    bsa = bsa.loc[
        bsa["DLY_Country"].fillna("").astype(str).str.len().le(3)
        & bsa["POR_Country"].fillna("").astype(str).str.len().le(3)
    ].copy()
    bsa["teu_bsa"] = to_number(bsa["TEU_BSA (Actual)"])
    bsa = bsa.rename(
        columns={
            "POR_Country": "origin",
            "POR_PORT": "ori_port",
            "DLY_Country": "dest",
            "DLY_PORT": "dst_port",
        }
    )
    keys = ["YYYYMM", "team", "origin", "ori_port", "dest", "dst_port"]
    return bsa.groupby(keys, dropna=False)["teu_bsa"].sum().reset_index()


def allocate_bsa_by_booking(
    rows: pd.DataFrame,
    bsa: pd.DataFrame,
    route_keys: list[str],
    weight_col: str = "BSA배분기준TEU",
) -> pd.DataFrame:
    bsa_by_route = bsa.groupby(route_keys, dropna=False)["teu_bsa"].sum().reset_index()
    rows = rows.merge(bsa_by_route, on=route_keys, how="left")
    rows["route_bsa"] = rows["teu_bsa"].fillna(0.0)
    route_totals = rows.groupby(route_keys, dropna=False).agg(
        route_basis=(weight_col, "sum"), route_bsa=("route_bsa", "first")
    )
    rows = rows.join(route_totals[["route_basis"]], on=route_keys)
    rows["BSA"] = np.where(
        rows["route_basis"].gt(0),
        rows["route_bsa"] * rows[weight_col] / rows["route_basis"],
        0.0,
    )
    return rows.drop(columns=["teu_bsa", "route_bsa", "route_basis"])


def build_result_tables(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, float]]:
    print("aggregating shipper-route performance")
    df = df.copy()
    df["origin"] = df["POR_CTR_CD"]
    df["ori_port"] = df["POR_PLC_CD"]
    df["dest"] = df["DLY_CTR_CD"]
    df["dst_port"] = df["DLY_PLC_CD"]
    df["업체명"] = df["BKG_SHPR_CST_ENM"].replace("", "(미지정)")
    df["영업사원"] = df["Salesman_POR"].replace("", "(미지정)")
    df["화주등급(A+B / C+D)"] = df["grade"].replace("", "(미지정)")
    df["구간"] = (
        df["origin"].fillna("")
        + "/"
        + df["ori_port"].fillna("")
        + " -> "
        + df["dest"].fillna("")
        + "/"
        + df["dst_port"].fillna("")
    )
    df["전체부킹"] = df["fst"]
    df["WOS-3"] = np.where(df["is_w3"], df["fst"], 0.0)
    df["WOS-3 normal"] = np.where(df["is_w3"] & df["is_normal"], df["lst"], 0.0)
    df["BSA배분기준TEU"] = np.where(df["is_normal"], df["lst"], 0.0)

    keys = [
        "YYYYMM",
        "team",
        "업체명",
        "영업사원",
        "origin",
        "ori_port",
        "dest",
        "dst_port",
        "구간",
        "고수익화주 태그",
        "화주등급(A+B / C+D)",
    ]
    monthly = df.groupby(keys, dropna=False).agg(
        전체부킹=("전체부킹", "sum"),
        **{"WOS-3": ("WOS-3", "sum"), "WOS-3 normal": ("WOS-3 normal", "sum")},
        BSA배분기준TEU=("BSA배분기준TEU", "sum"),
    ).reset_index()

    bsa = load_bsa()
    monthly_route_keys = ["YYYYMM", "team", "origin", "ori_port", "dest", "dst_port"]
    monthly = allocate_bsa_by_booking(monthly, bsa, monthly_route_keys)

    display_cols = [
        "팀",
        "업체명",
        "영업사원",
        "POR CN",
        "POR Port",
        "DLY CN",
        "DLY Port",
        "BSA",
        "전체부킹",
        "WOS-3",
        "BSA대비 비중(WOS-3/BSA)",
        "WOS-3 normal",
        "실선적률(WOS-3 normal/WOS-3)",
        "고수익화주 태그",
        "화주등급(A+B / C+D)",
    ]
    monthly["BSA대비 비중(WOS-3/BSA)"] = np.where(monthly["BSA"].gt(0), monthly["WOS-3"] / monthly["BSA"], np.nan)
    monthly["실선적률(WOS-3 normal/WOS-3)"] = np.where(
        monthly["WOS-3"].gt(0), monthly["WOS-3 normal"] / monthly["WOS-3"], np.nan
    )
    monthly["월"] = monthly["YYYYMM"].str[:4] + "-" + monthly["YYYYMM"].str[4:]
    monthly = monthly.rename(columns={"team": "팀"})
    monthly["POR CN"] = monthly["origin"]
    monthly["POR Port"] = monthly["ori_port"]
    monthly["DLY CN"] = monthly["dest"]
    monthly["DLY Port"] = monthly["dst_port"]
    monthly_out = monthly[["월", *display_cols]].copy()

    total_keys = [
        "team",
        "업체명",
        "영업사원",
        "origin",
        "ori_port",
        "dest",
        "dst_port",
        "구간",
        "고수익화주 태그",
        "화주등급(A+B / C+D)",
    ]
    total = monthly.rename(columns={"팀": "team"}).groupby(total_keys, dropna=False).agg(
        전체부킹=("전체부킹", "sum"),
        **{"WOS-3": ("WOS-3", "sum"), "WOS-3 normal": ("WOS-3 normal", "sum")},
        BSA배분기준TEU=("BSA배분기준TEU", "sum"),
    ).reset_index()
    total = allocate_bsa_by_booking(total, bsa, ["team", "origin", "ori_port", "dest", "dst_port"])
    total = total.rename(columns={"team": "팀"})
    total["BSA대비 비중(WOS-3/BSA)"] = np.where(total["BSA"].gt(0), total["WOS-3"] / total["BSA"], np.nan)
    total["실선적률(WOS-3 normal/WOS-3)"] = np.where(
        total["WOS-3"].gt(0), total["WOS-3 normal"] / total["WOS-3"], np.nan
    )
    total["POR CN"] = total["origin"]
    total["POR Port"] = total["ori_port"]
    total["DLY CN"] = total["dest"]
    total["DLY Port"] = total["dst_port"]
    total_out = total[display_cols].copy()

    sort_monthly = ["월", "팀", "POR CN", "POR Port", "DLY CN", "DLY Port", "WOS-3", "업체명"]
    monthly_out = monthly_out.sort_values(
        sort_monthly,
        ascending=[True, True, True, True, True, True, False, True],
    )
    total_out = total_out.sort_values(
        ["팀", "POR CN", "POR Port", "DLY CN", "DLY Port", "WOS-3", "업체명"],
        ascending=[True, True, True, True, True, False, True],
    )

    for frame in [monthly_out, total_out]:
        frame["BSA"] = frame["BSA"].round(4)
        for col in ["전체부킹", "WOS-3", "WOS-3 normal"]:
            frame[col] = frame[col].round(2)
        for col in ["BSA대비 비중(WOS-3/BSA)", "실선적률(WOS-3 normal/WOS-3)"]:
            frame[col] = frame[col].replace([np.inf, -np.inf], np.nan)

    validation = {
        "monthly_rows": float(len(monthly_out)),
        "total_rows": float(len(total_out)),
        "allocated_bsa_sum": float(monthly_out["BSA"].sum()),
        "raw_bsa_sum": float(bsa["teu_bsa"].sum()),
        "wos3_sum": float(monthly_out["WOS-3"].sum()),
        "booking_sum": float(monthly_out["전체부킹"].sum()),
    }
    return total_out, monthly_out, validation


def write_xlsx(total: pd.DataFrame, monthly: pd.DataFrame) -> Path:
    OUT_DIR.mkdir(exist_ok=True)
    path = OUT_DIR / "2025_업체별_구간별_BSA_배분.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        total.to_excel(writer, sheet_name="2025 전체", index=False)
        monthly.to_excel(writer, sheet_name="월별", index=False)

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
        widths = [12, 34, 16, 12, 12, 12, 12, 14, 14, 14, 18, 16, 24, 16, 20, 20, 20]
        for idx, width in enumerate(widths, start=1):
            if idx <= ws.max_column:
                ws.column_dimensions[get_column_letter(idx)].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = ws.cell(1, cell.column).value
                if header in {"BSA대비 비중(WOS-3/BSA)", "실선적률(WOS-3 normal/WOS-3)"}:
                    cell.number_format = "0.0%"
                elif header == "BSA":
                    cell.number_format = "#,##0.0000"
                elif header in {"전체부킹", "WOS-3", "WOS-3 normal"}:
                    cell.number_format = "#,##0.00"
    wb.save(path)
    return path


def upload_as_google_sheet(xlsx_path: Path) -> dict[str, str]:
    try:
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaFileUpload
        from google.auth.transport.requests import Request
    except Exception as exc:
        return {"error": f"Google API library unavailable: {exc}"}

    creds_dir = ROOT.parent / ".gdrive-mcp"
    credentials_path = creds_dir / "credentials.json"
    token_path = creds_dir / "token.json"
    if not credentials_path.exists() or not token_path.exists():
        return {"error": "Google Drive credentials were not found"}

    installed = json.loads(credentials_path.read_text(encoding="utf-8-sig"))["installed"]
    token = json.loads(token_path.read_text(encoding="utf-8-sig"))
    creds = Credentials(
        token=token.get("access_token"),
        refresh_token=token["refresh_token"],
        token_uri=installed["token_uri"],
        client_id=installed["client_id"],
        client_secret=installed["client_secret"],
    )
    creds.refresh(Request())
    service = build("drive", "v3", credentials=creds)
    media = MediaFileUpload(
        str(xlsx_path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True,
    )
    folder_id = os.environ.get("GDRIVE_FOLDER_ID", "1JIxg6Y-_gRfI1HueXZ1Q9j4-Z5bxvNgv")
    body = {
        "name": "2025 업체별 구간별 BSA 배분",
        "mimeType": "application/vnd.google-apps.spreadsheet",
        "parents": [folder_id],
    }
    created = service.files().create(
        body=body,
        media_body=media,
        fields="id,name,webViewLink",
        supportsAllDrives=True,
    ).execute()
    return created


def main() -> None:
    booking = add_dashboard_fields(build_booking_frame())
    total, monthly, validation = build_result_tables(booking)
    xlsx_path = write_xlsx(total, monthly)
    upload_result = upload_as_google_sheet(xlsx_path)
    report_path = OUT_DIR / "2025_업체별_구간별_BSA_배분_report.json"
    report_path.write_text(
        json.dumps(
            {
                "xlsx_path": str(xlsx_path),
                "validation": validation,
                "google_sheet": upload_result,
                "bsa_allocation_rule": (
                    "BSA is allocated by total booking TEU, not by WOS-3 TEU. "
                    "The allocation basis is restricted to Normal performance "
                    "(LST_Status=Normal LST_TEU). The monthly sheet allocates each month-route BSA "
                    "by month-route Normal performance share; the yearly sheet allocates annual route "
                    "BSA by annual route Normal performance share."
                ),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"xlsx={xlsx_path}")
    print(f"report={report_path}")
    print(f"google_sheet={upload_result}")
    print(f"validation={validation}")


if __name__ == "__main__":
    main()
